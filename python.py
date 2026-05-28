from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import math

wb = Workbook()

# Color constants
FILL_TITLE = PatternFill("solid", fgColor="99CCFF")
FILL_USECASE_HEADER = PatternFill("solid", fgColor="E2EFD9")
FILL_COL_HEADER = PatternFill("solid", fgColor="CCFFFF")
FILL_ROW_ALT = PatternFill("solid", fgColor="EFEFEF")
FILL_CATEGORY = PatternFill("solid", fgColor="D9E1F2")
FILL_NONE = PatternFill("solid", fgColor="FFFFFF")

FONT_TITLE = Font(name="Arial", bold=True, size=12)
FONT_HEADER = Font(name="Arial", bold=True, size=10)
FONT_BOLD = Font(name="Arial", bold=True, size=10)
FONT_NORMAL = Font(name="Arial", size=10)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def set_cell(ws, row, col, value, font=None, fill=None, align=None, border=True):
    cell = ws.cell(row=row, column=col)
    cell.value = value
    if font: cell.font = font
    if fill: cell.fill = fill
    if align: cell.alignment = align
    if border: cell.border = thin_border
    return cell

def apply_style(cell, fill, align, font=FONT_NORMAL):
    cell.fill = fill
    cell.alignment = align
    cell.border = thin_border
    cell.font = font

def apply_border_to_range(ws, start_row, end_row, start_col, end_col):
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            ws.cell(row=r, column=c).border = thin_border

def calculate_row_height(texts_with_widths, font_size=10):
    max_lines = 1
    for text, width in texts_with_widths:
        if not text: continue
        # Heuristic: approx 1.2 characters per unit of width for Arial 10
        chars_per_line = max(1, width * 1.1)
        lines = len(str(text)) / chars_per_line
        # Count explicit newlines
        lines += str(text).count('\n')
        if lines > max_lines:
            max_lines = lines
    return max(15, math.ceil(max_lines) * (font_size + 4))

def build_sheet(wb, sheet_name, use_case_title, categories_data, sheet_type="Function"):
    ws = wb.create_sheet(sheet_name)
    
    # Column widths
    col_widths = [2, 18, 17, 32, 30, 28, 18, 35, 33, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Status Data Validation
    dv = DataValidation(type="list", formula1='"Passed,Failed,Not Run,Not Completed"', allow_blank=False)
    ws.add_data_validation(dv)

    # Row 1: Title
    ws.merge_cells("A1:J1")
    set_cell(ws, 1, 1, f"{sheet_type} Test Cases", FONT_TITLE, FILL_TITLE, ALIGN_CENTER)
    apply_border_to_range(ws, 1, 1, 1, 10)
    ws.row_dimensions[1].height = 22

    # Row 2: Use Case name & stats header
    ws.merge_cells("A2:G3")
    set_cell(ws, 2, 1, f"Use Case: {use_case_title}", FONT_BOLD, FILL_USECASE_HEADER, ALIGN_CENTER)
    apply_border_to_range(ws, 2, 3, 1, 7)

    # Stats formulas (placeholders, will update last_row at the end)
    set_cell(ws, 2, 8, "Passed", FONT_NORMAL, FILL_NONE, ALIGN_CENTER)
    set_cell(ws, 3, 8, "Failed", FONT_NORMAL, FILL_NONE, ALIGN_CENTER)
    set_cell(ws, 4, 8, "Not Run", FONT_NORMAL, FILL_NONE, ALIGN_CENTER)
    set_cell(ws, 5, 8, "Not Completed", FONT_NORMAL, FILL_NONE, ALIGN_CENTER)
    set_cell(ws, 6, 8, "Number of test cases", FONT_NORMAL, FILL_NONE, ALIGN_CENTER)

    # Row 7-8: Column headers
    ws.merge_cells("B7:B8")
    set_cell(ws, 7, 2, "Category", FONT_HEADER, FILL_COL_HEADER, ALIGN_CENTER)
    ws.merge_cells("C7:C8")
    set_cell(ws, 7, 3, "Test Case ID", FONT_HEADER, FILL_COL_HEADER, ALIGN_CENTER)
    ws.merge_cells("D7:D8")
    set_cell(ws, 7, 4, "Test Case Description", FONT_HEADER, FILL_COL_HEADER, ALIGN_CENTER)
    ws.merge_cells("E7:F7")
    set_cell(ws, 7, 5, "Test Procedures", FONT_HEADER, FILL_COL_HEADER, ALIGN_CENTER)
    ws.merge_cells("G7:G8")
    set_cell(ws, 7, 7, "Test Input Value", FONT_HEADER, FILL_COL_HEADER, ALIGN_CENTER)
    ws.merge_cells("H7:H8")
    set_cell(ws, 7, 8, "Test Case Expected Result", FONT_HEADER, FILL_COL_HEADER, ALIGN_CENTER)
    ws.merge_cells("I7:I8")
    set_cell(ws, 7, 9, "Test Case Program Result", FONT_HEADER, FILL_COL_HEADER, ALIGN_CENTER)
    ws.merge_cells("J7:J8")
    set_cell(ws, 7, 10, "Status", FONT_HEADER, FILL_COL_HEADER, ALIGN_CENTER)

    set_cell(ws, 8, 5, "Steps to Perform", FONT_HEADER, FILL_COL_HEADER, ALIGN_LEFT)
    set_cell(ws, 8, 6, "Step Expected Result", FONT_HEADER, FILL_COL_HEADER, ALIGN_LEFT)

    apply_border_to_range(ws, 7, 8, 2, 10)
    
    # Auto height for headers
    ws.row_dimensions[7].height = 25
    ws.row_dimensions[8].height = 25

    # Fill test case data
    current_row = 9
    total_tcs = 0
    for category_name, test_cases in categories_data:
        cat_start = current_row
        for tc_index, tc in enumerate(test_cases):
            steps = tc.get("steps", [])
            if not steps:
                continue
                
            total_tcs += 1
            tc_id = tc["id"]
            tc_desc = tc["desc"]
            expected_result = tc["expected"]
            tc_start = current_row
            tc_fill = FILL_ROW_ALT if (total_tcs % 2 == 0) else FILL_NONE

            for i, (step_text, step_expected, input_val) in enumerate(steps):
                # Apply style and value to all columns
                for col in range(2, 11):
                    cell = ws.cell(row=current_row, column=col)
                    apply_style(cell, tc_fill, ALIGN_CENTER)
                    
                # Specific values and alignments
                if i == 0:
                    ws.cell(row=current_row, column=3).value = tc_id
                    ws.cell(row=current_row, column=4).value = tc_desc
                    ws.cell(row=current_row, column=4).alignment = ALIGN_LEFT
                    ws.cell(row=current_row, column=8).value = expected_result
                    ws.cell(row=current_row, column=8).alignment = ALIGN_LEFT
                    ws.cell(row=current_row, column=9).value = ""
                    ws.cell(row=current_row, column=10).value = "Not Run"
                    dv.add(ws.cell(row=current_row, column=10))

                ws.cell(row=current_row, column=5).value = step_text
                ws.cell(row=current_row, column=5).alignment = ALIGN_LEFT
                ws.cell(row=current_row, column=6).value = step_expected
                ws.cell(row=current_row, column=6).alignment = ALIGN_LEFT
                ws.cell(row=current_row, column=7).value = input_val

                # Dynamic row height
                row_texts = [
                    (step_text, 30),
                    (step_expected, 28),
                    (tc_desc if i == 0 else "", 32),
                    (expected_result if i == 0 else "", 35)
                ]
                ws.row_dimensions[current_row].height = calculate_row_height(row_texts)
                
                current_row += 1

            # Merge and fix borders for TC columns
            tc_end = current_row - 1
            for col in [3, 4, 8, 9, 10]:
                if tc_start != tc_end:
                    ws.merge_cells(start_row=tc_start, start_column=col, end_row=tc_end, end_column=col)
                
                # Re-apply border and alignment to the merged range
                align = ALIGN_LEFT if col in [4, 8] else ALIGN_CENTER
                for r in range(tc_start, tc_end + 1):
                    c = ws.cell(row=r, column=col)
                    c.border = thin_border
                    c.alignment = align

        # Merge category column
        if current_row > cat_start:
            ws.merge_cells(start_row=cat_start, start_column=2, end_row=current_row-1, end_column=2)
            c = ws.cell(row=cat_start, column=2)
            c.value = category_name
            c.font = FONT_BOLD
            c.fill = FILL_CATEGORY
            c.alignment = ALIGN_CENTER
            apply_border_to_range(ws, cat_start, current_row - 1, 2, 2)

    last_row = current_row - 1
    
    # Update Stats formulas with dynamic range
    ws.merge_cells("H2:I2")
    set_cell(ws, 2, 10, f'=COUNTIF($J$9:$J${last_row},"Passed")', FONT_NORMAL, FILL_NONE, ALIGN_CENTER)
    ws.merge_cells("H3:I3")
    set_cell(ws, 3, 10, f'=COUNTIF($J$9:$J${last_row},"Failed")', FONT_NORMAL, FILL_NONE, ALIGN_CENTER)
    ws.merge_cells("H4:I4")
    set_cell(ws, 4, 10, f'=COUNTIF($J$9:$J${last_row},"Not Run")', FONT_NORMAL, FILL_NONE, ALIGN_CENTER)
    ws.merge_cells("H5:I5")
    set_cell(ws, 5, 10, f'=COUNTIF($J$9:$J${last_row},"Not Completed")', FONT_NORMAL, FILL_NONE, ALIGN_CENTER)
    ws.merge_cells("H6:I6")
    set_cell(ws, 6, 10, total_tcs, FONT_NORMAL, FILL_NONE, ALIGN_CENTER) # Using direct count
    
    apply_border_to_range(ws, 2, 6, 8, 10)

    # Auto filter
    ws.auto_filter.ref = f"B7:J{last_row}"

    # Freeze header rows
    ws.freeze_panes = "A9"
    return ws


# ===================== DATA =====================

# UC3.1 - Tiếp đón người đến khám (check in)
uc31_func = [
    {
        "id": "UC3.1_FUNC_001",
        "desc": "Check-in thành công cho bệnh nhân đã có lịch khám ở trạng thái CONFIRMED.",
        "steps": [
            ("1. Đăng nhập hệ thống với tài khoản Lễ tân", "Đăng nhập thành công, hiển thị giao diện chính", "letan01 / Matkhau@123"),
            ("2. Truy cập menu \"Tiếp đón bệnh nhân\"", "Hiển thị danh sách lịch khám trong ngày", ""),
            ("3. Nhập mã lịch khám hoặc SĐT bệnh nhân vào ô tìm kiếm", "Thông tin lịch khám đúng bệnh nhân được tìm thấy và trạng thái hiện tại là CONFIRMED", "LK00234 / 0901234567"),
            ("4. Nhấn nút \"Check-in\"", "Hệ thống xác nhận thao tác check-in thành công", ""),
        ],
        "expected": "Trạng thái lịch khám cập nhật thành CHECKED_IN. Hệ thống tự động ghi nhận thời gian check-in của bệnh nhân."
    },
    {
        "id": "UC3.1_FUNC_002",
        "desc": "Tìm kiếm bệnh nhân trong danh sách tiếp đón theo nhiều tiêu chí.",
        "steps": [
            ("1. Truy cập menu \"Tiếp đón bệnh nhân\"", "Hiển thị danh sách bệnh nhân chờ tiếp đón", ""),
            ("2. Nhập họ tên bệnh nhân cần tìm vào ô tìm kiếm", "Hiển thị danh sách bệnh nhân có tên tương ứng", "Trần Văn A"),
            ("3. Xóa từ khóa cũ, nhập số điện thoại bệnh nhân vào ô tìm kiếm", "Hiển thị bệnh nhân có số điện thoại tương ứng", "0987654321"),
        ],
        "expected": "Hệ thống tìm kiếm realtime và trả về kết quả chính xác, tốc độ tìm kiếm nhanh (dưới 2 giây)."
    },
    {
        "id": "UC3.1_FUNC_003",
        "desc": "Đăng ký khám trực tiếp (không hẹn trước) thành công.",
        "steps": [
            ("1. Truy cập menu \"Tiếp đón bệnh nhân\"", "Hiển thị danh sách tiếp đón", ""),
            ("2. Nhập tìm kiếm thông tin bệnh nhân nhưng không có kết quả", "Hiển thị danh sách rỗng", "0999888777"),
            ("3. Nhấn nút \"Đăng ký khám mới\"", "Hiển thị form nhập thông tin bệnh nhân và phân ca", ""),
            ("4. Nhập Họ và tên, SĐT, Ngày sinh của bệnh nhân", "", "Nguyễn Thị Mai / 0999888777 / 20/10/1995"),
            ("5. Chọn Bác sĩ đang trực và Ca khám còn trống", "", "BS. Trần Thị Bích / Ca Sáng"),
            ("6. Nhấn \"Lưu đăng ký\"", "Hệ thống hiển thị thông báo đăng ký khám trực tiếp thành công", ""),
        ],
        "expected": "Tạo thành công lịch khám mới với trạng thái CHECKED_IN. Bệnh nhân được thêm vào danh sách chờ khám của bác sĩ đã chọn."
    },
    {
        "id": "UC3.1_FUNC_004",
        "desc": "Tiếp đón bệnh nhân đến trễ ca khám so với giờ hẹn.",
        "steps": [
            ("1. Truy cập menu \"Tiếp đón bệnh nhân\"", "Hiển thị danh sách lịch khám", ""),
            ("2. Tìm kiếm bệnh nhân có giờ hẹn đã qua so với giờ hiện tại", "Hiển thị thông tin lịch khám", "LK00235"),
            ("3. Nhấn nút \"Check-in\"", "Hệ thống phát hiện quá giờ khám và hiển thị cảnh báo đến trễ", ""),
            ("4. Lễ tân chọn \"Tiếp tục check-in\"", "Hệ thống thực hiện check-in cho bệnh nhân", ""),
        ],
        "expected": "Hệ thống cho phép tiếp tục check-in sau khi xác nhận cảnh báo. Trạng thái lịch khám chuyển sang CHECKED_IN và lưu log đến trễ."
    },
    {
        "id": "UC3.1_FUNC_005",
        "desc": "Check-in lịch khám có trạng thái PENDING (chưa được CONFIRMED).",
        "steps": [
            ("1. Truy cập menu \"Tiếp đón bệnh nhân\"", "Hiển thị danh sách lịch khám", ""),
            ("2. Tìm kiếm lịch khám đang ở trạng thái PENDING", "Hiển thị thông tin lịch khám", "LK00236"),
            ("3. Cố gắng nhấn nút \"Check-in\"", "Hệ thống từ chối thực hiện thao tác", ""),
        ],
        "expected": "Hệ thống hiển thị thông báo lỗi: \"Chỉ lịch khám ở trạng thái CONFIRMED mới được phép check-in!\""
    },
    {
        "id": "UC3.1_FUNC_006",
        "desc": "Check-in cho lịch khám đã bị Hủy (CANCELLED).",
        "steps": [
            ("1. Truy cập menu \"Tiếp đón bệnh nhân\"", "Hiển thị danh sách lịch khám", ""),
            ("2. Tìm lịch khám đã bị hủy (CANCELLED)", "Hiển thị lịch khám với trạng thái CANCELLED", "LK00237"),
            ("3. Thao tác nút \"Check-in\" (nếu có hiển thị)", "Hệ thống chặn thao tác hoặc nút bị disable", ""),
        ],
        "expected": "Hệ thống không cho phép check-in đối với lịch khám đã bị hủy."
    },
    {
        "id": "UC3.1_FUNC_007",
        "desc": "Cố gắng check-in cho bệnh nhân đã được CHECKED_IN trước đó.",
        "steps": [
            ("1. Tìm kiếm bệnh nhân đã có trạng thái CHECKED_IN", "Hiển thị thông tin lịch khám", "LK00238"),
            ("2. Cố gắng thực hiện lại thao tác \"Check-in\"", "Hệ thống hiển thị cảnh báo bệnh nhân đã check-in", ""),
        ],
        "expected": "Hệ thống hiển thị thông báo \"Bệnh nhân đã check-in trước đó!\". Không làm thay đổi thông tin thời gian check-in ban đầu."
    },
]

uc31_categories = [
    ("UC3.1 Tiếp đón & Tìm kiếm", uc31_func[:2]),
    ("UC3.1 Đăng ký & Đến trễ", uc31_func[2:4]),
    ("UC3.1 Ràng buộc & Ngoại lệ", uc31_func[4:]),
]

# UC3.2 - Khám bệnh và cập nhật hồ sơ
uc32_func = [
    {
        "id": "UC3.2_FUNC_001",
        "desc": "Khám bệnh và cập nhật hồ sơ bệnh án, hoàn tất ca khám hợp lệ.",
        "steps": [
            ("1. Đăng nhập hệ thống với tài khoản Bác sĩ có ca trực", "Đăng nhập thành công, hiển thị giao diện bác sĩ", "bsbich01 / Matkhau@123"),
            ("2. Mở danh sách bệnh nhân chờ khám và chọn bệnh nhân đã CHECKED_IN", "Hiển thị giao diện khám bệnh và bệnh sử", "Nguyễn Văn A"),
            ("3. Nhập thông tin triệu chứng lâm sàng", "", "Đau nhức răng số 8 hàm dưới bên trái, nướu sưng đỏ"),
            ("4. Nhập thông tin Chẩn đoán", "", "Răng khôn mọc lệch đâm vào răng số 7"),
            ("5. Chỉ định dịch vụ điều trị từ danh mục dịch vụ", "Dịch vụ được thêm vào phiếu chỉ định dịch vụ", "Phẫu thuật nhổ răng khôn mọc lệch"),
            ("6. Nhấn nút \"Hoàn tất khám\"", "Hệ thống hiển thị thông báo hoàn thành ca khám thành công", ""),
        ],
        "expected": "Hồ sơ khám bệnh được lưu trữ thành công. Trạng thái lịch khám chuyển thành COMPLETED. Thông tin chuyển tiếp sang bộ phận Thanh toán."
    },
    {
        "id": "UC3.2_FUNC_002",
        "desc": "Thao tác lưu tạm hồ sơ khám khi chưa hoàn thành quá trình khám.",
        "steps": [
            ("1. Bác sĩ mở giao diện khám bệnh của bệnh nhân", "Hiển thị hồ sơ bệnh nhân", ""),
            ("2. Nhập một số thông tin triệu chứng lâm sàng sơ bộ", "", "Răng ê buốt khi ăn đồ lạnh"),
            ("3. Nhấn nút \"Lưu tạm\"", "Hệ thống hiển thị thông báo lưu tạm thành công", ""),
        ],
        "expected": "Hồ sơ được lưu tạm với trạng thái IN_PROGRESS. Bệnh nhân vẫn nằm trong danh sách đang khám của bác sĩ để tiếp tục cập nhật."
    },
    {
        "id": "UC3.2_FUNC_003",
        "desc": "Chuyển bệnh nhân sang Bác sĩ khác (do yêu cầu chuyên khoa sâu).",
        "steps": [
            ("1. Bác sĩ đang khám nhấn nút \"Chuyển bác sĩ\"", "Hiển thị danh sách bác sĩ đang trong ca trực", ""),
            ("2. Chọn bác sĩ thuộc chuyên khoa sâu phù hợp và nhập lý do", "Bác sĩ chuyên khoa Implant / Yêu cầu cấy ghép", "BS. Nguyễn Văn C"),
            ("3. Xác nhận thao tác chuyển", "Hệ thống thông báo chuyển bác sĩ thành công", ""),
        ],
        "expected": "Thông tin bác sĩ phụ trách ca khám được cập nhật thành BS. Nguyễn Văn C. Bệnh nhân tự động chuyển sang danh sách chờ khám của bác sĩ mới."
    },
    {
        "id": "UC3.2_FUNC_004",
        "desc": "Cố gắng hoàn tất ca khám nhưng bỏ trống trường Chẩn đoán bắt buộc.",
        "steps": [
            ("1. Bác sĩ mở hồ sơ khám bệnh, nhập triệu chứng lâm sàng", "", "Răng sứt mẻ nhẹ"),
            ("2. Chọn dịch vụ điều trị", "", "Trám răng Composite"),
            ("3. Bỏ trống trường Chẩn đoán", "", ""),
            ("4. Nhấn nút \"Hoàn tất khám\"", "Hệ thống ngăn chặn thao tác và hiển thị cảnh báo lỗi", ""),
        ],
        "expected": "Hệ thống hiển thị thông báo lỗi \"Vui lòng nhập Chẩn đoán trước khi hoàn tất khám!\". Không lưu hồ sơ khám ở trạng thái hoàn thành."
    },
    {
        "id": "UC3.2_FUNC_005",
        "desc": "Bác sĩ truy cập và cố gắng khám bệnh nhân ngoài ca trực hoặc không thuộc phân công.",
        "steps": [
            ("1. Đăng nhập hệ thống bằng tài khoản Bác sĩ ngoài ca trực", "Đăng nhập thành công", "bsanh01 / Matkhau@123"),
            ("2. Cố gắng chọn bệnh nhân trong danh sách chờ của ca trực hiện tại để khám", "Hệ thống từ chối quyền truy cập ca khám", ""),
        ],
        "expected": "Hệ thống hiển thị cảnh báo lỗi: \"Bạn không có ca trực trong khung giờ này hoặc không được phân công khám bệnh nhân này!\""
    },
    {
        "id": "UC3.2_FUNC_006",
        "desc": "Bác sĩ xem bệnh sử cũ và upload ảnh chụp X-ray răng vào bệnh án.",
        "steps": [
            ("1. Bác sĩ mở hồ sơ bệnh án hiện tại của bệnh nhân", "Hiển thị thông tin khám", ""),
            ("2. Nhấn tab \"Lịch sử khám\" để xem các đợt điều trị cũ", "Hiển thị danh sách các đợt khám và chẩn đoán trước đây", ""),
            ("3. Tại khu vực hình ảnh y khoa, nhấn \"Tải ảnh X-ray lên\"", "Hiển thị hộp thoại chọn tập tin từ máy tính", ""),
            ("4. Chọn tập tin hình ảnh chụp X-ray hợp lệ và nhấn \"Lưu\"", "Tải file thành công, ảnh X-ray hiển thị trong hồ sơ bệnh án", "xray_rang_so_8.png"),
        ],
        "expected": "Hồ sơ hiển thị đầy đủ và chính xác bệnh sử cũ. Ảnh X-ray được upload thành công và hiển thị rõ ràng trên bệnh án."
    },
]

uc32_categories = [
    ("UC3.2 Khám bệnh & Hoàn tất", uc32_func[:3]),
    ("UC3.2 Ràng buộc & Bệnh án", uc32_func[3:]),
]

# UC3.3 - Thanh toán chi phí khám bệnh
uc33_func = [
    {
        "id": "UC3.3_FUNC_001",
        "desc": "Thanh toán thành công hóa đơn khám bệnh bằng phương thức chuyển khoản QR ngân hàng.",
        "steps": [
            ("1. Đăng nhập hệ thống với tài khoản Lễ tân", "Đăng nhập thành công, hiển thị giao diện tiếp đón/thanh toán", "letan01 / Matkhau@123"),
            ("2. Truy cập menu \"Thanh toán hóa đơn\"", "Hiển thị danh sách hóa đơn chờ thanh toán", ""),
            ("3. Tìm kiếm bệnh nhân vừa hoàn tất khám (COMPLETED) theo họ tên hoặc SĐT", "Thông tin hóa đơn hiển thị với danh sách dịch vụ đã sử dụng và tổng tiền", "Nguyễn Văn A"),
            ("4. Chọn phương thức thanh toán \"Chuyển khoản ngân hàng (QR Code)\"", "Hệ thống tạo mã QR động chứa thông tin số tiền và nội dung chuyển khoản", ""),
            ("5. Xác nhận giao dịch chuyển khoản thành công và nhấn \"Thanh toán\"", "Hệ thống thông báo thanh toán thành công, hiển thị hộp thoại in hóa đơn", ""),
        ],
        "expected": "Trạng thái hóa đơn cập nhật thành PAID. Hệ thống ghi nhận doanh thu thực tế. Tạo file hóa đơn PDF đẹp để in cho khách hàng."
    },
    {
        "id": "UC3.3_FUNC_002",
        "desc": "Thanh toán hóa đơn có áp dụng mã giảm giá (Discount) hợp lệ.",
        "steps": [
            ("1. Lễ tân mở chi tiết hóa đơn của bệnh nhân chờ thanh toán", "Hiển thị thông tin chi tiết hóa đơn", "LK00234"),
            ("2. Nhập mã giảm giá hợp lệ vào ô Mã giảm giá", "Mã được chấp nhận, hệ thống tính lại số tiền giảm và tổng tiền mới", "MEC10 (Giảm 10%)"),
            ("3. Chọn phương thức thanh toán \"Tiền mặt\" và nhấn \"Thanh toán\"", "Hệ thống xử lý thanh toán thành công", ""),
        ],
        "expected": "Tổng tiền thanh toán giảm chính xác 10%. Hóa đơn lưu thông tin: tổng tiền gốc, số tiền giảm giá, và số tiền thực tế khách hàng đã trả."
    },
    {
        "id": "UC3.3_FUNC_003",
        "desc": "Thanh toán hóa đơn bằng hình thức kết hợp nhiều phương thức thanh toán (Tiền mặt và Thẻ).",
        "steps": [
            ("1. Lễ tân mở chi tiết hóa đơn có tổng trị giá thanh toán 2.000.000đ", "Hiển thị hóa đơn", ""),
            ("2. Chọn phương thức thanh toán \"Hỗn hợp\"", "Hiển thị các ô nhập số tiền cho từng phương thức thanh toán", ""),
            ("3. Nhập số tiền mặt thanh toán", "", "1.000.000"),
            ("4. Nhập số tiền thanh toán qua thẻ ngân hàng", "", "1.000.000"),
            ("5. Xác nhận giao dịch thành công trên POS và nhấn \"Thanh toán\"", "Hệ thống báo thanh toán thành công", ""),
        ],
        "expected": "Trạng thái hóa đơn chuyển sang PAID. Hệ thống ghi nhận chính xác 1.000.000đ tiền mặt và 1.000.000đ tiền thẻ trong doanh thu chi tiết."
    },
    {
        "id": "UC3.3_FUNC_004",
        "desc": "Thanh toán cho lịch khám của bệnh nhân chưa hoàn tất quy trình khám bệnh (không ở trạng thái COMPLETED).",
        "steps": [
            ("1. Lễ tân truy cập menu \"Thanh toán hóa đơn\"", "Hiển thị danh sách hóa đơn", ""),
            ("2. Tìm kiếm bệnh nhân có lịch khám đang ở trạng thái CHECKED_IN hoặc IN_PROGRESS", "Thông tin bệnh nhân hiển thị nhưng không có hóa đơn hợp lệ để thanh toán", "Lê Văn D"),
            ("3. Thao tác cố gắng tạo hóa đơn hoặc thanh toán (nếu hệ thống có lỗ hổng)", "Hệ thống từ chối xử lý và hiển thị thông báo lỗi", ""),
        ],
        "expected": "Hệ thống ngăn chặn thanh toán, hiển thị thông báo lỗi \"Bệnh nhân chưa hoàn tất khám bệnh, không thể thanh toán chi phí!\""
    },
    {
        "id": "UC3.3_FUNC_005",
        "desc": "Cố gắng chỉnh sửa hoặc thay đổi thông tin dịch vụ trên hóa đơn đã ở trạng thái đã thanh toán (PAID).",
        "steps": [
            ("1. Lễ tân tìm kiếm và mở một hóa đơn đã thanh toán thành công (PAID)", "Hiển thị chi tiết hóa đơn đã thanh toán", "HD00123"),
            ("2. Thử thực hiện thay đổi giá trị dịch vụ, thêm/xóa dịch vụ hoặc chỉnh sửa số tiền", "Các nút chỉnh sửa bị ẩn hoặc disable. Hệ thống chặn mọi thao tác sửa đổi", ""),
        ],
        "expected": "Hóa đơn đã thanh toán (PAID) bị khóa hoàn toàn, không thể chỉnh sửa thông tin dịch vụ hay số tiền dưới mọi hình thức."
    },
    {
        "id": "UC3.3_FUNC_006",
        "desc": "Ngăn chặn hành vi nhấn thanh toán liên tiếp nhiều lần gây trùng lặp giao dịch (Double Payment).",
        "steps": [
            ("1. Lễ tân nhấn nút \"Thanh toán\" của hóa đơn", "Giao dịch bắt đầu xử lý", ""),
            ("2. Trong khi hệ thống đang xử lý giao dịch, cố tình nhấn liên tục vào nút \"Thanh toán\"", "Nút \"Thanh toán\" tự động bị disable ngay lập tức và các click sau không hoạt động", ""),
        ],
        "expected": "Hệ thống chỉ xử lý giao dịch đầu tiên và tạo ra duy nhất một hóa đơn PAID. Không phát sinh giao dịch trùng lặp trong cơ sở dữ liệu."
    },
]

uc33_categories = [
    ("UC3.3 Quy trình thanh toán", uc33_func[:3]),
    ("UC3.3 Ràng buộc & Bảo mật", uc33_func[3:]),
]

# UC3.4 - Thống kê doanh thu
uc34_func = [
    {
        "id": "UC3.4_FUNC_001",
        "desc": "Thống kê doanh thu theo khoảng thời gian hợp lệ (ngày/tuần/tháng/năm).",
        "steps": [
            ("1. Đăng nhập hệ thống bằng tài khoản Quản lý hoặc Admin", "Đăng nhập thành công, hiển thị dashboard quản lý", "admin / Matkhau@123"),
            ("2. Truy cập menu \"Thống kê doanh thu\"", "Hiển thị giao diện báo cáo doanh thu và bộ lọc", ""),
            ("3. Chọn khoảng thời gian Từ ngày \"01/05/2026\" Đến ngày \"28/05/2026\"", "", "01/05/2026 - 28/05/2026"),
            ("4. Nhấn nút \"Thống kê\"", "Hệ thống thực hiện truy vấn và hiển thị dữ liệu doanh thu tương ứng", ""),
        ],
        "expected": "Hệ thống hiển thị bảng số liệu doanh thu chính xác, tổng doanh thu khớp hoàn toàn với tổng các hóa đơn trạng thái PAID trong khoảng thời gian đã chọn. Biểu đồ trực quan hiển thị đúng."
    },
    {
        "id": "UC3.4_FUNC_002",
        "desc": "Thống kê doanh thu lọc theo Bác sĩ phụ trách.",
        "steps": [
            ("1. Truy cập chức năng \"Thống kê doanh thu\"", "Hiển thị giao diện thống kê", ""),
            ("2. Tại bộ lọc Bác sĩ, chọn \"BS. Trần Thị Bích\" và chọn khoảng thời gian", "", "BS. Trần Thị Bích / Tháng 5/2026"),
            ("3. Nhấn nút \"Thống kê\"", "Hiển thị bảng chi tiết doanh thu và biểu đồ liên quan", ""),
        ],
        "expected": "Hệ thống hiển thị chính xác các hóa đơn và tổng doanh thu phát sinh từ các ca khám do \"BS. Trần Thị Bích\" thực hiện."
    },
    {
        "id": "UC3.4_FUNC_003",
        "desc": "Xuất báo cáo doanh thu ra tệp tin Excel / PDF phục vụ báo cáo.",
        "steps": [
            ("1. Thực hiện thao tác thống kê doanh thu theo bộ lọc mong muốn", "Bảng thống kê hiển thị số liệu", ""),
            ("2. Nhấn nút \"Xuất Excel\" hoặc \"Xuất PDF\"", "Hệ thống tạo và tải xuống tập tin báo cáo doanh thu", ""),
        ],
        "expected": "Tệp tin Excel / PDF được tải xuống thành công, hiển thị đầy đủ, chính xác bảng biểu và biểu đồ doanh thu trùng khớp hoàn toàn với giao diện hệ thống."
    },
    {
        "id": "UC3.4_FUNC_004",
        "desc": "Thống kê doanh thu với khoảng thời gian không hợp lệ (Ngày bắt đầu lớn hơn Ngày kết thúc).",
        "steps": [
            ("1. Chọn bộ lọc Từ ngày \"28/05/2026\" Đến ngày \"01/05/2026\"", "", "28/05/2026 - 01/05/2026"),
            ("2. Nhấn nút \"Thống kê\"", "Hệ thống ngăn chặn truy vấn và hiển thị thông báo lỗi thời gian", ""),
        ],
        "expected": "Hệ thống hiển thị cảnh báo lỗi \"Khoảng thời gian không hợp lệ. Ngày bắt đầu phải nhỏ hơn hoặc bằng ngày kết thúc!\". Không thực hiện tìm kiếm dữ liệu."
    },
    {
        "id": "UC3.4_FUNC_005",
        "desc": "Người dùng không có quyền Quản lý/Admin cố gắng truy cập chức năng thống kê doanh thu.",
        "steps": [
            ("1. Đăng nhập hệ thống bằng tài khoản Lễ tân hoặc Bác sĩ", "Đăng nhập thành công", "letan01 / Matkhau@123"),
            ("2. Thử truy cập trực tiếp đường dẫn URL của chức năng thống kê doanh thu", "Hệ thống chặn truy cập và hiển thị thông báo lỗi phân quyền", "/admin/statistics"),
        ],
        "expected": "Hệ thống từ chối truy cập và hiển thị thông báo: \"Bạn không có quyền truy cập chức năng thống kê doanh thu!\"."
    },
    {
        "id": "UC3.4_FUNC_006",
        "desc": "Xác nhận doanh thu chỉ được tính từ các hóa đơn đã được thanh toán thành công (PAID).",
        "steps": [
            ("1. Tạo một hóa đơn mới ở trạng thái COMPLETED (chưa thanh toán) trong khoảng thời gian thống kê", "Hóa đơn mới được lưu nhưng chưa thanh toán", ""),
            ("2. Thực hiện chức năng Thống kê doanh thu cho khoảng thời gian đó", "Hiển thị số liệu thống kê doanh thu", ""),
        ],
        "expected": "Doanh thu hiển thị không cộng thêm giá trị của hóa đơn chưa thanh toán. Doanh thu chỉ được tính dựa trên các hóa đơn có trạng thái PAID."
    },
]

uc34_categories = [
    ("UC3.4 Lọc & Thống kê", uc34_func[:3]),
    ("UC3.4 Phân quyền & Ràng buộc", uc34_func[3:]),
]

# Remove default sheet, build all sheets
if 'Sheet' in wb.sheetnames:
    del wb['Sheet']

# Build Functional Test sheets
build_sheet(wb, "UC3.1 - Tiếp đón bệnh nhân", "Tiếp đón người đến khám (Check-in)", uc31_categories, "Function")
build_sheet(wb, "UC3.2 - Khám bệnh", "Khám bệnh và cập nhật hồ sơ", uc32_categories, "Function")
build_sheet(wb, "UC3.3 - Thanh toán chi phí", "Thanh toán chi phí khám bệnh", uc33_categories, "Function")
build_sheet(wb, "UC3.4 - Thống kê doanh thu", "Thống kê doanh thu", uc34_categories, "Function")

output_path = "G:/Project/Kiểm thử/N01_TH1_NhaKhoa_TestCase.xlsx"
wb.save(output_path)
print("Done!")
