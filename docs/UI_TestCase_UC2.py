from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import math

wb = Workbook()

# Color constants
FILL_TITLE = PatternFill("solid", fgColor="99CCFF")
FILL_USECASE_HEADER = PatternFill("solid", fgColor="E2EFD9")
FILL_COL_HEADER = PatternFill("solid", fgColor="CCFFFF")
FILL_ROW_ALT = PatternFill("solid", fgColor="EFEFEF")
FILL_CATEGORY = PatternFill("solid", fgColor="D9E1F2")
FILL_NONE = PatternFill("solid", fgColor="FFFFFF")

FONT_TITLE = Font(name="Arial", bold=True, size=12)
FONT_HEADER = Font(name="Arial", bold=True, size=10)
FONT_BOLD = Font(name="Arial", bold=True, size=10)
FONT_NORMAL = Font(name="Arial", size=10)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def set_cell(ws, row, col, value, font=None, fill=None, align=None, border=True):
    cell = ws.cell(row=row, column=col)
    cell.value = value
    if font: cell.font = font
    if fill: cell.fill = fill
    if align: cell.alignment = align
    if border: cell.border = thin_border
    return cell

def apply_style(cell, fill, align, font=FONT_NORMAL):
    cell.fill = fill
    cell.alignment = align
    cell.border = thin_border
    cell.font = font

def apply_border_to_range(ws, start_row, end_row, start_col, end_col):
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            ws.cell(row=r, column=c).border = thin_border

def calculate_row_height(texts_with_widths, font_size=10):
    max_lines = 1
    for text, width in texts_with_widths:
        if not text: continue
        chars_per_line = max(1, width * 1.1)
        lines = len(str(text)) / chars_per_line
        lines += str(text).count('\n')
        if lines > max_lines:
            max_lines = lines
    return max(15, math.ceil(max_lines) * (font_size + 4))

def build_sheet(wb, sheet_name, use_case_title, categories_data, sheet_type="UI"):
    ws = wb.create_sheet(sheet_name)
    
    col_widths = [2, 18, 17, 32, 30, 28, 18, 35, 33, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    dv = DataValidation(type="list", formula1='"Passed,Failed,Not Run,Not Completed"', allow_blank=False)
    ws.add_data_validation(dv)

    ws.merge_cells("A1:J1")
    set_cell(ws, 1, 1, f"{sheet_type} Test Cases", FONT_TITLE, FILL_TITLE, ALIGN_CENTER)
    apply_border_to_range(ws, 1, 1, 1, 10)
    ws.row_dimensions[1].height = 22

    ws.merge_cells("A2:G3")
    set_cell(ws, 2, 1, f"Use Case: {use_case_title}", FONT_BOLD, FILL_USECASE_HEADER, ALIGN_CENTER)
    apply_border_to_range(ws, 2, 3, 1, 7)

    set_cell(ws, 2, 8, "Passed", FONT_NORMAL, FILL_NONE, ALIGN_CENTER)
    set_cell(ws, 3, 8, "Failed", FONT_NORMAL, FILL_NONE, ALIGN_CENTER)
    set_cell(ws, 4, 8, "Not Run", FONT_NORMAL, FILL_NONE, ALIGN_CENTER)
    set_cell(ws, 5, 8, "Not Completed", FONT_NORMAL, FILL_NONE, ALIGN_CENTER)
    set_cell(ws, 6, 8, "Number of test cases", FONT_NORMAL, FILL_NONE, ALIGN_CENTER)

    ws.merge_cells("B7:B8")
    set_cell(ws, 7, 2, "Category", FONT_HEADER, FILL_COL_HEADER, ALIGN_CENTER)
    ws.merge_cells("C7:C8")
    set_cell(ws, 7, 3, "Test Case ID", FONT_HEADER, FILL_COL_HEADER, ALIGN_CENTER)
    ws.merge_cells("D7:D8")
    set_cell(ws, 7, 4, "Test Case Description", FONT_HEADER, FILL_COL_HEADER, ALIGN_CENTER)
    ws.merge_cells("E7:F7")
    set_cell(ws, 7, 5, "Test Procedures", FONT_HEADER, FILL_COL_HEADER, ALIGN_CENTER)
    ws.merge_cells("G7:G8")
    set_cell(ws, 7, 7, "Test Input Value", FONT_HEADER, FILL_COL_HEADER, ALIGN_CENTER)
    ws.merge_cells("H7:H8")
    set_cell(ws, 7, 8, "Test Case Expected Result", FONT_HEADER, FILL_COL_HEADER, ALIGN_CENTER)
    ws.merge_cells("I7:I8")
    set_cell(ws, 7, 9, "Test Case Program Result", FONT_HEADER, FILL_COL_HEADER, ALIGN_CENTER)
    ws.merge_cells("J7:J8")
    set_cell(ws, 7, 10, "Status", FONT_HEADER, FILL_COL_HEADER, ALIGN_CENTER)

    set_cell(ws, 8, 5, "Steps to Perform", FONT_HEADER, FILL_COL_HEADER, ALIGN_LEFT)
    set_cell(ws, 8, 6, "Step Expected Result", FONT_HEADER, FILL_COL_HEADER, ALIGN_LEFT)

    apply_border_to_range(ws, 7, 8, 2, 10)
    
    ws.row_dimensions[7].height = 25
    ws.row_dimensions[8].height = 25

    current_row = 9
    total_tcs = 0
    for category_name, test_cases in categories_data:
        cat_start = current_row
        for tc in test_cases:
            steps = tc.get("steps", [])
            if not steps:
                continue
                
            total_tcs += 1
            tc_id = tc["id"]
            tc_desc = tc["desc"]
            expected_result = tc["expected"]
            tc_start = current_row
            tc_fill = FILL_ROW_ALT if (total_tcs % 2 == 0) else FILL_NONE

            for i, (step_text, step_expected, input_val) in enumerate(steps):
                for col in range(2, 11):
                    cell = ws.cell(row=current_row, column=col)
                    apply_style(cell, tc_fill, ALIGN_CENTER)
                    
                if i == 0:
                    ws.cell(row=current_row, column=3).value = tc_id
                    ws.cell(row=current_row, column=4).value = tc_desc
                    ws.cell(row=current_row, column=4).alignment = ALIGN_LEFT
                    ws.cell(row=current_row, column=8).value = expected_result
                    ws.cell(row=current_row, column=8).alignment = ALIGN_LEFT
                    ws.cell(row=current_row, column=9).value = ""
                    ws.cell(row=current_row, column=10).value = "Not Run"
                    dv.add(ws.cell(row=current_row, column=10))

                ws.cell(row=current_row, column=5).value = step_text
                ws.cell(row=current_row, column=5).alignment = ALIGN_LEFT
                ws.cell(row=current_row, column=6).value = step_expected
                ws.cell(row=current_row, column=6).alignment = ALIGN_LEFT
                ws.cell(row=current_row, column=7).value = input_val

                row_texts = [
                    (step_text, 30),
                    (step_expected, 28),
                    (tc_desc if i == 0 else "", 32),
                    (expected_result if i == 0 else "", 35)
                ]
                ws.row_dimensions[current_row].height = calculate_row_height(row_texts)
                
                current_row += 1

            tc_end = current_row - 1
            for col in [3, 4, 8, 9, 10]:
                if tc_start != tc_end:
                    ws.merge_cells(start_row=tc_start, start_column=col, end_row=tc_end, end_column=col)
                
                align = ALIGN_LEFT if col in [4, 8] else ALIGN_CENTER
                for r in range(tc_start, tc_end + 1):
                    c = ws.cell(row=r, column=col)
                    c.border = thin_border
                    c.alignment = align

        if current_row > cat_start:
            ws.merge_cells(start_row=cat_start, start_column=2, end_row=current_row-1, end_column=2)
            c = ws.cell(row=cat_start, column=2)
            c.value = category_name
            c.font = FONT_BOLD
            c.fill = FILL_CATEGORY
            c.alignment = ALIGN_CENTER
            apply_border_to_range(ws, cat_start, current_row - 1, 2, 2)

    last_row = current_row - 1
    
    ws.merge_cells("H2:I2")
    set_cell(ws, 2, 10, f'=COUNTIF($J$9:$J${last_row},"Passed")', FONT_NORMAL, FILL_NONE, ALIGN_CENTER)
    ws.merge_cells("H3:I3")
    set_cell(ws, 3, 10, f'=COUNTIF($J$9:$J${last_row},"Failed")', FONT_NORMAL, FILL_NONE, ALIGN_CENTER)
    ws.merge_cells("H4:I4")
    set_cell(ws, 4, 10, f'=COUNTIF($J$9:$J${last_row},"Not Run")', FONT_NORMAL, FILL_NONE, ALIGN_CENTER)
    ws.merge_cells("H5:I5")
    set_cell(ws, 5, 10, f'=COUNTIF($J$9:$J${last_row},"Not Completed")', FONT_NORMAL, FILL_NONE, ALIGN_CENTER)
    ws.merge_cells("H6:I6")
    set_cell(ws, 6, 10, total_tcs, FONT_NORMAL, FILL_NONE, ALIGN_CENTER)
    
    apply_border_to_range(ws, 2, 6, 8, 10)
    ws.auto_filter.ref = f"B7:J{last_row}"
    ws.freeze_panes = "A9"
    return ws


# ===================== UI TEST CASES DATA FOR UC2 =====================

# UC2.1 Thiết lập các ngày nghỉ - UI Test Cases
uc21_ui = [
    {
        "id": "UC2.1_UI_001",
        "desc": "Kiểm tra hiển thị giao diện thiết lập ngày nghỉ và bảng danh sách.",
        "steps": [
            ("1. Đăng nhập hệ thống với tài khoản Admin/Quản lý", "Đăng nhập thành công, hiển thị trang dashboard quản trị", "admin / Matkhau@123"),
            ("2. Truy cập chức năng \"Thiết lập các ngày nghỉ\" từ menu", "Hiển thị trang quản lý ngày nghỉ, danh sách các ngày nghỉ hiện có và nút \"Thêm mới\"", ""),
            ("3. Kiểm tra các cột hiển thị trong bảng", "Bảng hiển thị đầy đủ các cột: STT, Tên ngày nghỉ, Ngày bắt đầu, Ngày kết thúc, Loại nghỉ, Trạng thái, Thao tác", ""),
            ("4. Kiểm tra nút \"Thêm mới\"", "Nút hiển thị rõ ràng ở góc trên bên phải bảng, màu xanh lá hoặc xanh dương, có icon thêm", "")
        ],
        "expected": "Giao diện danh sách ngày nghỉ hiển thị đúng layout, font chữ Arial đồng bộ và responsive tốt trên màn hình tablet/laptop."
    },
    {
        "id": "UC2.1_UI_002",
        "desc": "Kiểm tra validation lỗi nhập liệu form Thêm mới ngày nghỉ.",
        "steps": [
            ("1. Click nút \"Thêm mới\" ngày nghỉ", "Form nhập thông tin hiển thị dưới dạng modal popup ở giữa màn hình, nền phía sau bị làm mờ", ""),
            ("2. Để trống các trường bắt buộc (Tên ngày nghỉ, Ngày bắt đầu) và click \"Lưu\"", "Hệ thống chặn lưu, các trường bỏ trống hiển thị viền đỏ (CSS :invalid) và xuất hiện thông báo lỗi ngay dưới trường", ""),
            ("3. Nhập Ngày kết thúc nhỏ hơn Ngày bắt đầu và click \"Lưu\"", "Trường Ngày kết thúc hiển thị cảnh báo lỗi định dạng/logic ngày bắt đầu lớn hơn ngày kết thúc", "Từ: 10/06/2026, Đến: 05/06/2026")
        ],
        "expected": "Form modal hiển thị các cảnh báo lỗi validation rõ ràng màu đỏ trực tiếp dưới các trường tương ứng."
    },
    {
        "id": "UC2.1_UI_003",
        "desc": "Kiểm tra hiển thị màu sắc trạng thái (badge) ngày nghỉ.",
        "steps": [
            ("1. Quan sát cột Trạng thái trên bảng danh sách ngày nghỉ", "Các trạng thái ACTIVE, INACTIVE được hiển thị bằng các nhãn (badge) có màu sắc khác nhau", ""),
            ("2. Kiểm tra màu sắc nhãn ACTIVE", "Hiển thị nhãn màu xanh lá cây với chữ \"Đang áp dụng\" hoặc \"ACTIVE\"", ""),
            ("3. Kiểm tra màu sắc nhãn INACTIVE", "Hiển thị nhãn màu xám với chữ \"Ngưng áp dụng\" hoặc \"INACTIVE\"", "")
        ],
        "expected": "Các nhãn trạng thái hiển thị đẹp mắt, độ tương phản cao giúp người quản trị dễ dàng phân biệt trạng thái của ngày nghỉ."
    }
]

# UC2.2 Thiết lập ca làm việc - UI Test Cases
uc22_ui = [
    {
        "id": "UC2.2_UI_001",
        "desc": "Kiểm tra hiển thị giao diện Quản lý ca làm việc và danh sách ca.",
        "steps": [
            ("1. Đăng nhập Quản lý và chọn \"Quản lý ca làm việc\"", "Giao diện danh sách các ca làm việc hiện có (ca sáng, ca chiều, ca tối) và nút \"Thêm ca\" hiển thị", "quanly01 / Matkhau@123"),
            ("2. Kiểm tra các cột trong bảng danh sách ca", "Bảng hiển thị đầy đủ các cột: STT, Tên ca, Giờ bắt đầu, Giờ kết thúc, Giới hạn bệnh nhân, Trạng thái, Thao tác", "")
        ],
        "expected": "Bảng danh sách ca làm việc hiển thị đúng layout, font chữ Arial rõ ràng, thông tin thời gian hiển thị đúng định dạng HH:MM."
    },
    {
        "id": "UC2.2_UI_002",
        "desc": "Kiểm tra validation lỗi nhập liệu form Thêm mới ca làm việc.",
        "steps": [
            ("1. Click nút \"Thêm ca\"", "Form nhập thông tin ca làm việc hiển thị dưới dạng modal popup ở giữa màn hình", ""),
            ("2. Nhập Giờ kết thúc nhỏ hơn Giờ bắt đầu và nhấn \"Lưu\"", "Hệ thống ngăn chặn thao tác lưu, trường Giờ kết thúc hiển thị cảnh báo lỗi và viền đỏ", "Bắt đầu: 13:00, Kết thúc: 11:00"),
            ("3. Nhập Giới hạn bệnh nhân nhỏ hơn hoặc bằng 0", "Hệ thống báo lỗi validation tại ô nhập Giới hạn bệnh nhân: \"Giới hạn bệnh nhân phải lớn hơn 0!\"", "Giới hạn: 0")
        ],
        "expected": "Giao diện hiển thị cảnh báo lỗi validation rõ ràng, ngăn chặn người dùng lưu dữ liệu không hợp lệ."
    },
    {
        "id": "UC2.2_UI_003",
        "desc": "Kiểm tra trạng thái disabled của nút ngưng hoạt động khi ca trực đang được sử dụng.",
        "steps": [
            ("1. Định vị ca làm việc đang có lịch trực bác sĩ hoặc lịch khám trong tương lai", "Trạng thái ca làm việc hiển thị ACTIVE và có liên kết dữ liệu", "Ca sáng"),
            ("2. Hover chuột vào nút \"Ngưng hoạt động\" hoặc \"Xóa\" của ca làm việc này", "Nút bấm hiển thị ở trạng thái disabled (mờ đi, con trỏ cursor dạng not-allowed) hoặc hiển thị tooltip cảnh báo lý do", "")
        ],
        "expected": "Nút thao tác bị vô hiệu hóa trực quan trên giao diện để tránh người quản lý ngưng ca trực làm ảnh hưởng dữ liệu đang vận hành."
    }
]

# UC2.3 Đăng ký lịch trực của bác sĩ - UI Test Cases
uc23_ui = [
    {
        "id": "UC2.3_UI_001",
        "desc": "Kiểm tra hiển thị giao diện danh sách lịch trực bác sĩ và form Đăng ký.",
        "steps": [
            ("1. Đăng nhập Quản lý và truy cập chức năng \"Lịch trực bác sĩ\"", "Màn hình hiển thị danh sách lịch trực (dạng danh sách hoặc dạng lịch calendar)", "quanly01 / Matkhau@123"),
            ("2. Click nút \"Đăng ký lịch trực\"", "Form đăng ký hiển thị dưới dạng modal popup với các dropdown chọn Bác sĩ, Ngày trực, Ca trực, Chuyên khoa", ""),
            ("3. Kiểm tra các dropdown lựa chọn", "Các dropdown hiển thị danh sách bác sĩ đang ACTIVE, các ca làm việc đang ACTIVE và các chuyên khoa", "")
        ],
        "expected": "Form modal hiển thị cân đối, các trường nhập liệu sắp xếp khoa học, dropdown hỗ trợ tìm kiếm nhanh."
    },
    {
        "id": "UC2.3_UI_002",
        "desc": "Kiểm tra validation cảnh báo khi đăng ký lịch trực trùng hoặc vào ngày nghỉ.",
        "steps": [
            ("1. Chọn bác sĩ và ngày/ca trực đã có lịch trực trước đó và nhấn \"Lưu\"", "Hệ thống từ chối lưu và hiển thị Toast message cảnh báo: \"Bác sĩ đã có lịch trực trong cùng thời gian!\"", "BS. Nguyễn Văn A / 05/06/2026"),
            ("2. Chọn ngày trực trùng với ngày nghỉ ACTIVE của phòng khám", "Hệ thống hiển thị viền đỏ ở ô chọn ngày và báo lỗi: \"Không thể đăng ký lịch trực vào ngày nghỉ!\"", "Ngày trực: 02/09/2026 (Ngày lễ)")
        ],
        "expected": "Hệ thống hiển thị thông báo lỗi rõ ràng, trực quan giúp người quản lý tránh thao tác đăng ký sai lệch."
    }
]

# UC2.4 Đăng ký lịch khám của bệnh nhân - UI Test Cases
uc24_ui = [
    {
        "id": "UC2.4_UI_001",
        "desc": "Kiểm tra hiển thị giao diện Đăng ký lịch khám và liên kết dữ liệu bác sĩ.",
        "steps": [
            ("1. Truy cập chức năng \"Đăng ký lịch khám\"", "Giao diện form đặt lịch hiển thị với các trường nhập thông tin bệnh nhân, chọn Ngày khám, Ca khám, Bác sĩ, Dịch vụ", ""),
            ("2. Chọn Ngày khám và Ca khám cụ thể", "Dữ liệu được thiết lập", "Ngày khám: 05/06/2026 / Ca sáng"),
            ("3. Click mở dropdown chọn Bác sĩ", "Dropdown bác sĩ chỉ hiển thị danh sách các bác sĩ có lịch trực ACTIVE trong ca khám đã chọn", "")
        ],
        "expected": "Giao diện liên kết dữ liệu tự động chính xác theo thời gian thực, ngăn chặn chọn bác sĩ không có ca trực."
    },
    {
        "id": "UC2.4_UI_002",
        "desc": "Kiểm tra validation khi full slot khám hoặc trùng lịch khám bệnh nhân.",
        "steps": [
            ("1. Chọn ngày và ca khám đã đạt giới hạn tối đa bệnh nhân", "Ca khám hiển thị nhãn \"Đầy lịch\" (màu đỏ/mờ) và không cho phép click chọn", "Ca sáng ngày 05/06/2026"),
            ("2. Nhập thông tin bệnh nhân đã có lịch khám khác trong cùng ca khám đó và click \"Đặt lịch\"", "Hệ thống hiển thị cảnh báo lỗi trùng lịch khám bệnh nhân: \"Bệnh nhân đã có lịch khám trong cùng thời gian!\"", "SĐT: 0987654321")
        ],
        "expected": "Giao diện hiển thị các cảnh báo lỗi validation rõ ràng, chặn gửi yêu cầu đặt lịch không hợp lệ lên máy chủ."
    }
]

# UC2.5 Theo dõi lịch khám - UI Test Cases
uc25_ui = [
    {
        "id": "UC2.5_UI_001",
        "desc": "Kiểm tra hiển thị giao diện Theo dõi lịch khám và các bộ lọc.",
        "steps": [
            ("1. Truy cập chức năng \"Theo dõi lịch khám\"", "Hiển thị danh sách lịch khám trong ngày hiện tại, thanh tìm kiếm bệnh nhân và bộ lọc trạng thái", ""),
            ("2. Kiểm tra các bộ lọc trên giao diện", "Hiển thị dropdown lọc theo Bác sĩ, Ca khám, Trạng thái (PENDING, CONFIRMED, CHECKED_IN, COMPLETED, CANCELLED, NO_SHOW)", ""),
            ("3. Kiểm tra ô tìm kiếm nhanh", "Ô tìm kiếm nhanh hiển thị placeholder: \"Tìm kiếm theo tên bệnh nhân hoặc mã lịch khám...\"", "")
        ],
        "expected": "Giao diện theo dõi lịch khám hiển thị đúng layout, thông tin bảng rõ ràng và dễ dàng thao tác lọc/tìm kiếm."
    },
    {
        "id": "UC2.5_UI_002",
        "desc": "Kiểm tra cập nhật nhanh trạng thái lịch khám trên giao diện.",
        "steps": [
            ("1. Chọn một lịch khám có trạng thái CONFIRMED và click nút \"Check-in\"", "Trạng thái lịch khám cập nhật ngay lập tức thành CHECKED_IN (màu xanh dương)", "LK00234"),
            ("2. Chọn lịch khám đang CHECKED_IN và click nút \"Hoàn thành khám\"", "Trạng thái lịch khám cập nhật thành COMPLETED (màu xanh lá)", "LK00234"),
            ("3. Chọn một lịch khám và click nút \"Không đến\" (No Show)", "Trạng thái lịch khám cập nhật thành NO_SHOW (màu xám nhạt)", "LK00235")
        ],
        "expected": "Trạng thái lịch khám cập nhật realtime trên bảng danh sách mà không cần load lại trang, màu sắc nhãn hiển thị chính xác."
    }
]

# UC2.6 Quản lý Bệnh nhân - UI Test Cases
uc26_ui = [
    {
        "id": "UC2.6_UI_001",
        "desc": "Kiểm tra hiển thị giao diện Quản lý bệnh nhân và nút Thêm mới.",
        "steps": [
            ("1. Đăng nhập hệ thống với tài khoản Lễ tân", "Đăng nhập thành công", "letan01 / Matkhau@123"),
            ("2. Truy cập chức năng \"Quản lý bệnh nhân\" từ menu", "Hiển thị bảng danh sách bệnh nhân và nút \"Thêm mới\"", ""),
            ("3. Kiểm tra các cột trong bảng danh sách bệnh nhân", "Bảng hiển thị đủ các cột: STT, Mã bệnh nhân, Họ tên, Ngày sinh, Giới tính, Số điện thoại, Địa chỉ, Trạng thái, Thao tác", "")
        ],
        "expected": "Giao diện danh sách bệnh nhân hiển thị đúng thiết kế, font chữ Arial rõ ràng và căn chỉnh chuẩn xác."
    },
    {
        "id": "UC2.6_UI_002",
        "desc": "Kiểm tra validation lỗi nhập liệu form Thêm mới/Cập nhật thông tin bệnh nhân.",
        "steps": [
            ("1. Click chọn nút \"Thêm mới\"", "Form popup modal thêm bệnh nhân hiển thị", ""),
            ("2. Để trống Họ tên và Số điện thoại, nhấn \"Lưu\"", "Hệ thống hiển thị viền đỏ báo lỗi tại các ô bắt buộc bỏ trống", ""),
            ("3. Nhập Ngày sinh lớn hơn ngày hiện tại hoặc Số điện thoại có định dạng chữ cái", "Hệ thống báo lỗi logic: \"Ngày sinh không được lớn hơn ngày hiện tại\" và \"Số điện thoại không hợp lệ\"", "Ngày sinh: 12/12/2030 / SĐT: 098abc123")
        ],
        "expected": "Modal hiển thị các lỗi validation trực quan màu đỏ hướng dẫn người dùng chỉnh sửa thông tin cho đúng."
    },
    {
        "id": "UC2.6_UI_003",
        "desc": "Kiểm tra chức năng tìm kiếm nhanh bệnh nhân.",
        "steps": [
            ("1. Nhập từ khóa tìm kiếm (Họ tên, Mã bệnh nhân hoặc SĐT) vào ô tìm kiếm", "Bảng danh sách bệnh nhân tự động lọc realtime kết quả khớp từ khóa", "SĐT: 0905123456"),
            ("2. Xóa trắng ô tìm kiếm", "Bảng danh sách khôi phục hiển thị đầy đủ danh sách bệnh nhân", "")
        ],
        "expected": "Hệ thống thực hiện tìm kiếm nhanh chóng, kết quả hiển thị chính xác theo bộ lọc từ khóa nhập vào."
    },
    {
        "id": "UC2.6_UI_004",
        "desc": "Kiểm tra màu sắc trạng thái (badge) bệnh nhân ACTIVE/INACTIVE.",
        "steps": [
            ("1. Quan sát cột Trạng thái trên bảng danh sách bệnh nhân", "Các bệnh nhân được gắn nhãn trạng thái tương ứng", ""),
            ("2. Kiểm tra màu sắc nhãn ACTIVE và INACTIVE", "Trạng thái ACTIVE hiển thị màu xanh lá (Hoạt động), INACTIVE hiển thị màu xám (Ngưng hoạt động)", "")
        ],
        "expected": "Màu sắc nhãn tương phản tốt, giúp phân biệt trạng thái của bệnh nhân nhanh chóng."
    }
]

uc21_categories = [
    ("UC2.1 Giao diện & Thêm mới", uc21_ui[:2]),
    ("UC2.1 Trạng thái ngày nghỉ", uc21_ui[2:]),
]

uc22_categories = [
    ("UC2.2 Giao diện ca trực", uc22_ui[:2]),
    ("UC2.2 Ngưng hoạt động ca", uc22_ui[2:]),
]

uc23_categories = [
    ("UC2.3 Giao diện & Đăng ký", uc23_ui),
]

uc24_categories = [
    ("UC2.4 Form & Liên kết", uc24_ui[:1]),
    ("UC2.4 Ràng buộc & Trùng", uc24_ui[1:]),
]

uc25_categories = [
    ("UC2.5 Giao diện & Bộ lọc", uc25_ui[:1]),
    ("UC2.5 Cập nhật trạng thái", uc25_ui[1:]),
]

uc26_categories = [
    ("UC2.6 Giao diện & Thêm mới", uc26_ui[:2]),
    ("UC2.6 Tìm kiếm & Trạng thái", uc26_ui[2:]),
]

# Remove default sheet, build all sheets
if 'Sheet' in wb.sheetnames:
    del wb['Sheet']

# Build UI Test sheets for UC2
build_sheet(wb, "UC2.1 - Thiết lập ngày nghỉ", "Thiết lập các ngày nghỉ phòng khám", uc21_categories, "UI")
build_sheet(wb, "UC2.2 - Thiết lập ca trực", "Thiết lập ca làm việc", uc22_categories, "UI")
build_sheet(wb, "UC2.3 - Đăng ký lịch trực", "Đăng ký lịch trực của bác sĩ", uc23_categories, "UI")
build_sheet(wb, "UC2.4 - Đăng ký lịch khám", "Đăng ký lịch khám của bệnh nhân", uc24_categories, "UI")
build_sheet(wb, "UC2.5 - Theo dõi lịch khám", "Theo dõi lịch khám bệnh nhân", uc25_categories, "UI")
build_sheet(wb, "UC2.6 - Quản lý bệnh nhân", "Quản lý Bệnh nhân (UI)", uc26_categories, "UI")

output_path = "G:/Project/Kiểm thử/UI_TestCase_UC2.xlsx"
wb.save(output_path)
print("Done creating UI_TestCase_UC2.xlsx!")
