from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import math

wb = Workbook()

# Color constants
FILL_TITLE = PatternFill("solid", fgColor="99CCFF")
FILL_USECASE_HEADER = PatternFill("solid", fgColor="E2EFD9")
FILL_COL_HEADER = PatternFill("solid", fgColor="CCFFFF")
FILL_ROW_ALT = PatternFill("solid", fgColor="EFEFEF")
FILL_CATEGORY = PatternFill("solid", fgColor="D9E1F2")
FILL_NONE = PatternFill("solid", fgColor="FFFFFF")

FONT_TITLE = Font(name="Arial", bold=True, size=12)
FONT_HEADER = Font(name="Arial", bold=True, size=10)
FONT_BOLD = Font(name="Arial", bold=True, size=10)
FONT_NORMAL = Font(name="Arial", size=10)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def set_cell(ws, row, col, value, font=None, fill=None, align=None, border=True):
    cell = ws.cell(row=row, column=col)
    cell.value = value
    if font: cell.font = font
    if fill: cell.fill = fill
    if align: cell.alignment = align
    if border: cell.border = thin_border
    return cell

def apply_style(cell, fill, align, font=FONT_NORMAL):
    cell.fill = fill
    cell.alignment = align
    cell.border = thin_border
    cell.font = font

def apply_border_to_range(ws, start_row, end_row, start_col, end_col):
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            ws.cell(row=r, column=c).border = thin_border

def calculate_row_height(texts_with_widths, font_size=10):
    max_lines = 1
    for text, width in texts_with_widths:
        if not text: continue
        chars_per_line = max(1, width * 1.1)
        lines = len(str(text)) / chars_per_line
        lines += str(text).count('\n')
        if lines > max_lines:
            max_lines = lines
    return max(15, math.ceil(max_lines) * (font_size + 4))

def build_sheet(wb, sheet_name, use_case_title, categories_data, sheet_type="Function"):
    ws = wb.create_sheet(sheet_name)
    
    col_widths = [2, 18, 17, 32, 30, 28, 18, 35, 33, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    dv = DataValidation(type="list", formula1='"Passed,Failed,Not Run,Not Completed"', allow_blank=False)
    ws.add_data_validation(dv)

    ws.merge_cells("A1:J1")
    set_cell(ws, 1, 1, f"{sheet_type} Test Cases", FONT_TITLE, FILL_TITLE, ALIGN_CENTER)
    apply_border_to_range(ws, 1, 1, 1, 10)
    ws.row_dimensions[1].height = 22

    ws.merge_cells("A2:G3")
    set_cell(ws, 2, 1, f"Use Case: {use_case_title}", FONT_BOLD, FILL_USECASE_HEADER, ALIGN_CENTER)
    apply_border_to_range(ws, 2, 3, 1, 7)

    set_cell(ws, 2, 8, "Passed", FONT_NORMAL, FILL_NONE, ALIGN_CENTER)
    set_cell(ws, 3, 8, "Failed", FONT_NORMAL, FILL_NONE, ALIGN_CENTER)
    set_cell(ws, 4, 8, "Not Run", FONT_NORMAL, FILL_NONE, ALIGN_CENTER)
    set_cell(ws, 5, 8, "Not Completed", FONT_NORMAL, FILL_NONE, ALIGN_CENTER)
    set_cell(ws, 6, 8, "Number of test cases", FONT_NORMAL, FILL_NONE, ALIGN_CENTER)

    ws.merge_cells("B7:B8")
    set_cell(ws, 7, 2, "Category", FONT_HEADER, FILL_COL_HEADER, ALIGN_CENTER)
    ws.merge_cells("C7:C8")
    set_cell(ws, 7, 3, "Test Case ID", FONT_HEADER, FILL_COL_HEADER, ALIGN_CENTER)
    ws.merge_cells("D7:D8")
    set_cell(ws, 7, 4, "Test Case Description", FONT_HEADER, FILL_COL_HEADER, ALIGN_CENTER)
    ws.merge_cells("E7:F7")
    set_cell(ws, 7, 5, "Test Procedures", FONT_HEADER, FILL_COL_HEADER, ALIGN_CENTER)
    ws.merge_cells("G7:G8")
    set_cell(ws, 7, 7, "Test Input Value", FONT_HEADER, FILL_COL_HEADER, ALIGN_CENTER)
    ws.merge_cells("H7:H8")
    set_cell(ws, 7, 8, "Test Case Expected Result", FONT_HEADER, FILL_COL_HEADER, ALIGN_CENTER)
    ws.merge_cells("I7:I8")
    set_cell(ws, 7, 9, "Test Case Program Result", FONT_HEADER, FILL_COL_HEADER, ALIGN_CENTER)
    ws.merge_cells("J7:J8")
    set_cell(ws, 7, 10, "Status", FONT_HEADER, FILL_COL_HEADER, ALIGN_CENTER)

    set_cell(ws, 8, 5, "Steps to Perform", FONT_HEADER, FILL_COL_HEADER, ALIGN_LEFT)
    set_cell(ws, 8, 6, "Step Expected Result", FONT_HEADER, FILL_COL_HEADER, ALIGN_LEFT)

    apply_border_to_range(ws, 7, 8, 2, 10)
    
    ws.row_dimensions[7].height = 25
    ws.row_dimensions[8].height = 25

    current_row = 9
    total_tcs = 0
    for category_name, test_cases in categories_data:
        cat_start = current_row
        for tc in test_cases:
            steps = tc.get("steps", [])
            if not steps:
                continue
                
            total_tcs += 1
            tc_id = tc["id"]
            tc_desc = tc["desc"]
            expected_result = tc["expected"]
            tc_start = current_row
            tc_fill = FILL_ROW_ALT if (total_tcs % 2 == 0) else FILL_NONE

            for i, (step_text, step_expected, input_val) in enumerate(steps):
                for col in range(2, 11):
                    cell = ws.cell(row=current_row, column=col)
                    apply_style(cell, tc_fill, ALIGN_CENTER)
                    
                if i == 0:
                    ws.cell(row=current_row, column=3).value = tc_id
                    ws.cell(row=current_row, column=4).value = tc_desc
                    ws.cell(row=current_row, column=4).alignment = ALIGN_LEFT
                    ws.cell(row=current_row, column=8).value = expected_result
                    ws.cell(row=current_row, column=8).alignment = ALIGN_LEFT
                    ws.cell(row=current_row, column=9).value = ""
                    ws.cell(row=current_row, column=10).value = "Not Run"
                    dv.add(ws.cell(row=current_row, column=10))

                ws.cell(row=current_row, column=5).value = step_text
                ws.cell(row=current_row, column=5).alignment = ALIGN_LEFT
                ws.cell(row=current_row, column=6).value = step_expected
                ws.cell(row=current_row, column=6).alignment = ALIGN_LEFT
                ws.cell(row=current_row, column=7).value = input_val

                row_texts = [
                    (step_text, 30),
                    (step_expected, 28),
                    (tc_desc if i == 0 else "", 32),
                    (expected_result if i == 0 else "", 35)
                ]
                ws.row_dimensions[current_row].height = calculate_row_height(row_texts)
                
                current_row += 1

            tc_end = current_row - 1
            for col in [3, 4, 8, 9, 10]:
                if tc_start != tc_end:
                    ws.merge_cells(start_row=tc_start, start_column=col, end_row=tc_end, end_column=col)
                
                align = ALIGN_LEFT if col in [4, 8] else ALIGN_CENTER
                for r in range(tc_start, tc_end + 1):
                    c = ws.cell(row=r, column=col)
                    c.border = thin_border
                    c.alignment = align

        if current_row > cat_start:
            ws.merge_cells(start_row=cat_start, start_column=2, end_row=current_row-1, end_column=2)
            c = ws.cell(row=cat_start, column=2)
            c.value = category_name
            c.font = FONT_BOLD
            c.fill = FILL_CATEGORY
            c.alignment = ALIGN_CENTER
            apply_border_to_range(ws, cat_start, current_row - 1, 2, 2)

    last_row = current_row - 1
    
    ws.merge_cells("H2:I2")
    set_cell(ws, 2, 10, f'=COUNTIF($J$9:$J${last_row},"Passed")', FONT_NORMAL, FILL_NONE, ALIGN_CENTER)
    ws.merge_cells("H3:I3")
    set_cell(ws, 3, 10, f'=COUNTIF($J$9:$J${last_row},"Failed")', FONT_NORMAL, FILL_NONE, ALIGN_CENTER)
    ws.merge_cells("H4:I4")
    set_cell(ws, 4, 10, f'=COUNTIF($J$9:$J${last_row},"Not Run")', FONT_NORMAL, FILL_NONE, ALIGN_CENTER)
    ws.merge_cells("H5:I5")
    set_cell(ws, 5, 10, f'=COUNTIF($J$9:$J${last_row},"Not Completed")', FONT_NORMAL, FILL_NONE, ALIGN_CENTER)
    ws.merge_cells("H6:I6")
    set_cell(ws, 6, 10, total_tcs, FONT_NORMAL, FILL_NONE, ALIGN_CENTER)
    
    apply_border_to_range(ws, 2, 6, 8, 10)
    ws.auto_filter.ref = f"B7:J{last_row}"
    ws.freeze_panes = "A9"
    return ws


# ===================== UC1 DATA =====================

# UC1.1 - Quản lý người dùng và phân quyền
uc11_func = [
    {
        "id": "UC1.1_FUNC_001",
        "desc": "Thêm mới tài khoản người dùng hợp lệ thành công với quyền Lễ tân.",
        "steps": [
            ("1. Đăng nhập hệ thống bằng tài khoản Admin", "Đăng nhập thành công, hiển thị trang quản trị", "admin / Matkhau@123"),
            ("2. Truy cập chức năng \"Quản lý người dùng\"", "Hiển thị danh sách tài khoản hiện có", ""),
            ("3. Chọn nút \"Thêm tài khoản\"", "Hiển thị form nhập thông tin tài khoản mới", ""),
            ("4. Nhập Tên đăng nhập, Mật khẩu hợp lệ, Họ tên, Vai trò, Số điện thoại và Email", "", "letan_test / Matkhau@123 / Lễ tân / Nguyễn Thị Letan / 0981234567 / letan@clinic.com"),
            ("5. Nhấn nút \"Lưu\"", "Hệ thống thông báo tạo tài khoản thành công và ghi nhận audit log", "")
        ],
        "expected": "Tài khoản mới được tạo thành công với trạng thái ACTIVE và phân quyền chính xác theo vai trò Lễ tân."
    },
    {
        "id": "UC1.1_FUNC_002",
        "desc": "Thêm tài khoản thất bại do trùng tên đăng nhập (Username).",
        "steps": [
            ("1. Truy cập form \"Thêm tài khoản\"", "Form nhập hiển thị", ""),
            ("2. Nhập thông tin tài khoản mới nhưng sử dụng Username trùng với tài khoản hiện có", "", "letan_test / Matkhau@123 / Lễ tân / Nguyễn Văn Trùng / 0987778889 / letan2@clinic.com"),
            ("3. Nhấn nút \"Lưu\"", "Hệ thống hiển thị lỗi báo trùng username và từ chối lưu dữ liệu", "")
        ],
        "expected": "Hệ thống hiển thị thông báo: \"Tên đăng nhập đã tồn tại trong hệ thống!\" và không tạo tài khoản mới."
    },
    {
        "id": "UC1.1_FUNC_003",
        "desc": "Thêm tài khoản thất bại do mật khẩu không đủ độ mạnh.",
        "steps": [
            ("1. Truy cập form \"Thêm tài khoản\"", "Form nhập hiển thị", ""),
            ("2. Nhập thông tin tài khoản mới với mật khẩu chỉ có 6 ký tự hoặc không có chữ hoa/số", "", "letan_weak / 123456 / Lễ tân / Bùi Văn Yếu / 0987654321 / weak@clinic.com"),
            ("3. Nhấn nút \"Lưu\"", "Hệ thống báo lỗi validation mật khẩu và không cho phép lưu", "")
        ],
        "expected": "Thông báo lỗi: \"Mật khẩu tối thiểu phải 8 ký tự, bao gồm chữ hoa, chữ thường và chữ số!\""
    },
    {
        "id": "UC1.1_FUNC_004",
        "desc": "Thêm tài khoản thất bại do thiếu thông tin bắt buộc.",
        "steps": [
            ("1. Truy cập form \"Thêm tài khoản\"", "Form nhập hiển thị", ""),
            ("2. Để trống trường \"Họ tên\" và \"Vai trò\", nhập các trường khác", "", "letan_empty / Matkhau@123 / (Trống) / (Trống) / 0987654321 / empty@clinic.com"),
            ("3. Nhấn nút \"Lưu\"", "Hệ thống báo lỗi thiếu thông tin bắt buộc và ngăn chặn lưu", "")
        ],
        "expected": "Hiển thị thông báo lỗi yêu cầu nhập đầy đủ thông tin bắt buộc."
    },
    {
        "id": "UC1.1_FUNC_005",
        "desc": "Chỉnh sửa thông tin và vai trò của tài khoản thành công.",
        "steps": [
            ("1. Chọn một tài khoản từ danh sách người dùng", "Thông tin tài khoản hiển thị", "letan_test"),
            ("2. Chọn nút \"Chỉnh sửa\"", "Form chỉnh sửa hiển thị thông tin hiện tại", ""),
            ("3. Cập nhật Số điện thoại mới và đổi vai trò thành Bác sĩ", "", "SĐT mới: 0912999888 / Vai trò: Bác sĩ"),
            ("4. Nhấn \"Lưu\"", "Hệ thống báo cập nhật thành công và ghi audit log", "")
        ],
        "expected": "Thông tin tài khoản và vai trò được cập nhật mới trên hệ thống. Tài khoản đổi sang quyền Bác sĩ."
    },
    {
        "id": "UC1.1_FUNC_006",
        "desc": "Khóa tài khoản người dùng đang hoạt động (chuyển sang INACTIVE).",
        "steps": [
            ("1. Chọn tài khoản đang ở trạng thái ACTIVE", "Chọn tài khoản cần khóa", "letan_test"),
            ("2. Nhấn nút \"Khóa tài khoản\" và xác nhận tại hộp thoại", "Hộp thoại yêu cầu xác nhận hiển thị, người dùng click Đồng ý", ""),
            ("3. Kiểm tra trạng thái của tài khoản trong danh sách", "Trạng thái hiển thị là INACTIVE", "")
        ],
        "expected": "Tài khoản bị khóa thành công, chuyển sang trạng thái INACTIVE và không thể đăng nhập vào hệ thống."
    },
    {
        "id": "UC1.1_FUNC_007",
        "desc": "Người dùng không có quyền quản trị cố gắng truy cập chức năng.",
        "steps": [
            ("1. Đăng nhập hệ thống bằng tài khoản Bác sĩ hoặc Lễ tân", "Đăng nhập thành công", "letan01 / Matkhau@123"),
            ("2. Cố gắng truy cập menu \"Quản lý người dùng\" hoặc gọi trực tiếp API quản lý tài khoản", "Hệ thống từ chối truy cập", "")
        ],
        "expected": "Hệ thống hiển thị thông báo lỗi phân quyền: \"Bạn không có quyền truy cập chức năng này!\""
    }
]

uc11_categories = [
    ("UC1.1 Tạo mới tài khoản", uc11_func[:4]),
    ("UC1.1 Sửa đổi & Khóa", uc11_func[4:6]),
    ("UC1.1 Phân quyền & Bảo mật", uc11_func[6:]),
]

# UC1.2 - Quản lý bác sĩ
uc12_func = [
    {
        "id": "UC1.2_FUNC_001",
        "desc": "Thêm mới hồ sơ bác sĩ hợp lệ thành công.",
        "steps": [
            ("1. Đăng nhập bằng tài khoản Admin/Quản lý", "Đăng nhập thành công, hiển thị trang dashboard", "quanly01 / Matkhau@123"),
            ("2. Truy cập chức năng \"Quản lý bác sĩ\"", "Hiển thị danh sách bác sĩ", ""),
            ("3. Chọn nút \"Thêm bác sĩ\"", "Hiển thị form nhập thông tin hồ sơ bác sĩ", ""),
            ("4. Nhập đầy đủ thông tin: Họ tên, Ngày sinh, Giới tính, SĐT, Email, Địa chỉ, Chuyên khoa, Số CCHN", "", "BS. Trần Quốc Bảo / 15/08/1985 / Nam / 0905123456 / baotran@clinic.com / 123 Nguyễn Trãi, Q5 / Răng Hàm Mặt / CCHN-001234"),
            ("5. Nhấn nút \"Lưu\"", "Hệ thống kiểm tra hợp lệ và thông báo tạo thành công", "")
        ],
        "expected": "Hồ sơ bác sĩ được lưu vào hệ thống ở trạng thái ACTIVE, liên kết với tài khoản hệ thống tương ứng và hiển thị trong danh sách đặt lịch."
    },
    {
        "id": "UC1.2_FUNC_002",
        "desc": "Thêm mới bác sĩ thất bại do trùng Số chứng chỉ hành nghề (CCHN).",
        "steps": [
            ("1. Truy cập form \"Thêm bác sĩ\"", "Form nhập hồ sơ hiển thị", ""),
            ("2. Nhập thông tin bác sĩ mới nhưng sử dụng Số CCHN đã tồn tại trong hệ thống", "", "BS. Lê Thị CCHN / 10/10/1990 / Nữ / 0905999888 / le_cchn@clinic.com / Đà Nẵng / Nội khoa / CCHN-001234 (Trùng)"),
            ("3. Nhấn nút \"Lưu\"", "Hệ thống báo lỗi trùng số CCHN và từ chối lưu", "")
        ],
        "expected": "Thông báo hiển thị: \"Số chứng chỉ hành nghề đã tồn tại!\" và không lưu hồ sơ mới."
    },
    {
        "id": "UC1.2_FUNC_003",
        "desc": "Thêm mới bác sĩ thất bại do thiếu thông tin bắt buộc.",
        "steps": [
            ("1. Truy cập form \"Thêm bác sĩ\"", "Form nhập hồ sơ hiển thị", ""),
            ("2. Để trống trường \"Họ tên\" và \"Số điện thoại\", nhập các trường khác hợp lệ", "", "(Trống) / 01/01/1990 / Nam / (Trống) / email@clinic.com / CCHN-99999"),
            ("3. Nhấn nút \"Lưu\"", "Hệ thống báo lỗi thiếu thông tin bắt buộc và không lưu", "")
        ],
        "expected": "Hệ thống hiển thị lỗi báo đỏ tại các trường Họ tên, Số điện thoại và yêu cầu nhập lại."
    },
    {
        "id": "UC1.2_FUNC_004",
        "desc": "Thêm mới bác sĩ thất bại do định dạng Email hoặc SĐT không hợp lệ.",
        "steps": [
            ("1. Truy cập form \"Thêm bác sĩ\"", "Form nhập hồ sơ hiển thị", ""),
            ("2. Nhập số điện thoại sai định dạng (chỉ có 5 số) và Email không có ký tự '@'", "", "BS. Nguyễn Định Dạng / 09055 / abc.com / CCHN-54321"),
            ("3. Nhấn nút \"Lưu\"", "Hệ thống kiểm tra dữ liệu và báo lỗi định dạng trường", "")
        ],
        "expected": "Hệ thống ngăn chặn lưu và báo lỗi: \"Số điện thoại hoặc Email không đúng định dạng!\""
    },
    {
        "id": "UC1.2_FUNC_005",
        "desc": "Xem chi tiết hồ sơ và cập nhật thông tin bác sĩ thành công.",
        "steps": [
            ("1. Chọn bác sĩ cần chỉnh sửa từ danh sách", "Hệ thống hiển thị đầy đủ thông tin hồ sơ bác sĩ", "BS. Trần Quốc Bảo"),
            ("2. Chọn nút \"Chỉnh sửa\"", "Các trường thông tin chuyển sang trạng thái cho phép nhập", ""),
            ("3. Thay đổi thông tin Địa chỉ, Bằng cấp/Kinh nghiệm", "", "Địa chỉ mới: 456 Lê Lợi, Q1 / Kinh nghiệm: 12 năm kinh nghiệm Răng Hàm Mặt"),
            ("4. Nhấn nút \"Lưu\"", "Hệ thống báo cập nhật thành công và ghi nhận lịch sử chỉnh sửa", "")
        ],
        "expected": "Thông tin bác sĩ được lưu cập nhật mới. Hồ sơ bác sĩ hiển thị đúng địa chỉ và kinh nghiệm mới."
    },
    {
        "id": "UC1.2_FUNC_006",
        "desc": "Ngưng hoạt động bác sĩ (chuyển sang trạng thái INACTIVE) thành công.",
        "steps": [
            ("1. Chọn bác sĩ trong danh sách bác sĩ", "Hiển thị thông tin bác sĩ", "BS. Trần Quốc Bảo"),
            ("2. Nhấn nút \"Ngưng hoạt động\" và xác nhận", "Hệ thống hỏi xác nhận, người dùng chọn Đồng ý", ""),
            ("3. Kiểm tra trạng thái bác sĩ trên danh sách đặt lịch khám", "Bác sĩ không còn xuất hiện trên giao diện đặt lịch khám", "")
        ],
        "expected": "Hồ sơ bác sĩ chuyển sang trạng thái INACTIVE. Bác sĩ bị ẩn khỏi danh sách đặt lịch hẹn khám bệnh."
    }
]

uc12_categories = [
    ("UC1.2 Tạo hồ sơ bác sĩ", uc12_func[:4]),
    ("UC1.2 Sửa đổi & Ngưng hoạt động", uc12_func[4:]),
]

# UC1.3 - Quản lý danh mục dịch vụ
uc13_func = [
    {
        "id": "UC1.3_FUNC_001",
        "desc": "Thêm mới dịch vụ nha khoa hợp lệ thành công.",
        "steps": [
            ("1. Đăng nhập bằng tài khoản Quản lý", "Đăng nhập thành công, hiển thị trang quản lý", "quanly01 / Matkhau@123"),
            ("2. Truy cập chức năng \"Quản lý danh mục dịch vụ\"", "Hiển thị danh sách dịch vụ hiện có của phòng khám", ""),
            ("3. Chọn nút \"Thêm dịch vụ\"", "Hiển thị form nhập thông tin dịch vụ", ""),
            ("4. Nhập Tên dịch vụ, Nhóm dịch vụ, Giá dịch vụ, Thời gian thực hiện và Mô tả", "", "Hàn răng sâu thẩm mỹ / Răng sứ thẩm mỹ / 500.000 / 30 phút / Hàn răng công nghệ mới"),
            ("5. Nhấn nút \"Lưu\"", "Hệ thống kiểm tra dữ liệu và báo lưu thành công", "")
        ],
        "expected": "Dịch vụ mới được tạo thành công với Service ID duy nhất, trạng thái ACTIVE và xuất hiện trong hệ thống đặt lịch khám, hóa đơn."
    },
    {
        "id": "UC1.3_FUNC_002",
        "desc": "Thêm dịch vụ thất bại do trùng tên dịch vụ trong cùng một nhóm.",
        "steps": [
            ("1. Truy cập form \"Thêm dịch vụ\"", "Form nhập hiển thị", ""),
            ("2. Nhập thông tin dịch vụ mới nhưng Tên dịch vụ trùng với dịch vụ đã tồn tại trong nhóm đó", "", "Hàn răng sâu thẩm mỹ / Răng sứ thẩm mỹ / 600.000 / 40 phút"),
            ("3. Nhấn nút \"Lưu\"", "Hệ thống từ chối lưu và hiển thị thông báo lỗi trùng tên", "")
        ],
        "expected": "Hiển thị thông báo: \"Tên dịch vụ đã tồn tại trong nhóm dịch vụ này!\" và không lưu dữ liệu."
    },
    {
        "id": "UC1.3_FUNC_003",
        "desc": "Thêm dịch vụ thất bại do giá hoặc thời gian thực hiện không hợp lệ.",
        "steps": [
            ("1. Truy cập form \"Thêm dịch vụ\"", "Form nhập hiển thị", ""),
            ("2. Nhập Giá dịch vụ < 0 hoặc Thời gian thực hiện <= 0", "", "Giá: -100.000 / Thời gian: 0 phút"),
            ("3. Nhấn nút \"Lưu\"", "Hệ thống báo lỗi validation giá trị âm/không phù hợp", "")
        ],
        "expected": "Thông báo lỗi: \"Giá dịch vụ phải >= 0 và thời gian thực hiện phải > 0 phút!\". Dịch vụ không được lưu."
    },
    {
        "id": "UC1.3_FUNC_004",
        "desc": "Chỉnh sửa thông tin dịch vụ thành công.",
        "steps": [
            ("1. Chọn dịch vụ cần chỉnh sửa từ danh sách", "Hiển thị thông tin dịch vụ", "Hàn răng sâu thẩm mỹ"),
            ("2. Cập nhật thông tin Giá dịch vụ và Thời gian thực hiện", "", "Giá mới: 550.000 / Thời gian mới: 45 phút"),
            ("3. Nhấn nút \"Lưu\"", "Hệ thống thông báo cập nhật thành công và ghi nhận lịch sử thay đổi", "")
        ],
        "expected": "Dữ liệu dịch vụ được cập nhật thành công. Giá mới được áp dụng cho các hóa đơn/lịch khám mới mà không ảnh hưởng hóa đơn cũ."
    },
    {
        "id": "UC1.3_FUNC_005",
        "desc": "Ngưng sử dụng dịch vụ (chuyển sang trạng thái INACTIVE).",
        "steps": [
            ("1. Chọn dịch vụ đang hoạt động trong danh sách", "Chọn dịch vụ cần ngưng", "Hàn răng sâu thẩm mỹ"),
            ("2. Nhấn nút \"Ngưng sử dụng\" và xác nhận", "Hệ thống yêu cầu xác nhận và chuyển trạng thái", ""),
            ("3. Kiểm tra sự xuất hiện của dịch vụ khi tạo lịch khám hoặc lập hóa đơn mới", "Dịch vụ không còn xuất hiện trong ô chọn dịch vụ", "")
        ],
        "expected": "Trạng thái dịch vụ chuyển sang INACTIVE và bị ẩn khỏi hệ thống đặt lịch khám/hóa đơn."
    }
]

uc13_categories = [
    ("UC1.3 Khai báo dịch vụ", uc13_func[:3]),
    ("UC1.3 Sửa đổi & Ngưng dùng", uc13_func[3:]),
]

# UC1.4 - Thiết lập giá dịch vụ
uc14_func = [
    {
        "id": "UC1.4_FUNC_001",
        "desc": "Thiết lập giá mới cho một dịch vụ thành công.",
        "steps": [
            ("1. Đăng nhập bằng tài khoản Quản lý", "Đăng nhập thành công, hiển thị trang quản trị", "quanly01 / Matkhau@123"),
            ("2. Truy cập chức năng \"Thiết lập giá dịch vụ\"", "Hiển thị danh sách dịch vụ hiện có của phòng khám", ""),
            ("3. Chọn dịch vụ cần thiết lập giá mới", "Hiển thị chi tiết giá hiện tại và lịch sử giá dịch vụ", "Dịch vụ: Hàn răng sâu"),
            ("4. Nhập mức giá mới, ngày hiệu lực và nhấn \"Lưu\"", "", "Giá mới: 600.000 / Ngày hiệu lực: Ngày mai"),
            ("5. Hệ thống kiểm tra dữ liệu và báo cập nhật thành công", "Thông báo thành công, ghi nhận lịch sử thay đổi giá", "")
        ],
        "expected": "Giá dịch vụ mới được lưu vào hệ thống, ghi nhận ngày hiệu lực. Giá mới sẽ tự động được áp dụng từ ngày hiệu lực trở đi."
    },
    {
        "id": "UC1.4_FUNC_002",
        "desc": "Thiết lập giá dịch vụ thất bại do mức giá nhỏ hơn hoặc bằng 0.",
        "steps": [
            ("1. Chọn dịch vụ và click cập nhật giá", "Hiển thị form nhập giá", ""),
            ("2. Nhập giá dịch vụ bằng 0 hoặc số âm", "", "Giá mới: 0 (hoặc -50.000)"),
            ("3. Nhấn nút \"Lưu\"", "Hệ thống báo lỗi mức giá không hợp lệ và từ chối lưu", "")
        ],
        "expected": "Thông báo lỗi: \"Giá dịch vụ phải lớn hơn 0!\". Hệ thống không lưu cấu hình mới."
    },
    {
        "id": "UC1.4_FUNC_003",
        "desc": "Thiết lập giá dịch vụ thất bại do trùng khoảng thời gian hiệu lực.",
        "steps": [
            ("1. Chọn dịch vụ đã có giá đang áp dụng", "Hiển thị thông tin dịch vụ", "Dịch vụ: Hàn răng sâu"),
            ("2. Thêm mức giá mới với ngày hiệu lực trùng hoặc nằm trong khoảng hiệu lực của mức giá ACTIVE khác", "", "Giá mới: 650.000 / Ngày hiệu lực: Trùng ngày hiệu lực đang có"),
            ("3. Nhấn nút \"Lưu\"", "Hệ thống từ chối lưu và cảnh báo trùng khoảng thời gian", "")
        ],
        "expected": "Cảnh báo lỗi: \"Trùng khoảng thời gian hiệu lực của giá dịch vụ!\". Không cho phép lưu."
    },
    {
        "id": "UC1.4_FUNC_004",
        "desc": "Cập nhật giá dịch vụ hàng loạt thành công.",
        "steps": [
            ("1. Chọn nhiều dịch vụ trong danh sách thiết lập giá", "Các dịch vụ được chọn hiển thị dấu tick", "Dịch vụ A, Dịch vụ B"),
            ("2. Nhập tỷ lệ phần trăm tăng/giảm giá và ngày áp dụng", "", "Tăng: 10% / Ngày hiệu lực: Đầu tháng sau"),
            ("3. Nhấn nút \"Cập nhật hàng loạt\"", "Hệ thống tự động tính toán giá mới và thông báo cập nhật thành công", "")
        ],
        "expected": "Toàn bộ dịch vụ được chọn được tạo mức giá mới với ngày hiệu lực tương ứng chính xác đến đơn vị đồng."
    },
    {
        "id": "UC1.4_FUNC_005",
        "desc": "Xem lịch sử thay đổi giá dịch vụ.",
        "steps": [
            ("1. Chọn một dịch vụ từ danh sách dịch vụ", "Hiển thị chi tiết dịch vụ", "Dịch vụ: Hàn răng sâu"),
            ("2. Chọn chức năng \"Xem lịch sử giá\"", "Hệ thống hiển thị danh sách các lần thay đổi giá trong quá khứ", "")
        ],
        "expected": "Hiển thị đầy đủ lịch sử gồm: Mức giá, Ngày bắt đầu hiệu lực, Ngày kết thúc hiệu lực (nếu có), Người thực hiện thay đổi."
    }
]

uc14_categories = [
    ("UC1.4 Thiết lập giá đơn lẻ", uc14_func[:3]),
    ("UC1.4 Thiết lập hàng loạt & Lịch sử", uc14_func[3:]),
]


# Remove default sheet, build all sheets
if 'Sheet' in wb.sheetnames:
    del wb['Sheet']

# Build Functional Test sheets for UC1
build_sheet(wb, "UC1.1 - Quản lý tài khoản", "Quản lý người dùng và phân quyền (RBAC)", uc11_categories, "Function")
build_sheet(wb, "UC1.2 - Quản lý bác sĩ", "Quản lý thông tin hồ sơ bác sĩ", uc12_categories, "Function")
build_sheet(wb, "UC1.3 - Danh mục dịch vụ", "Quản lý danh mục dịch vụ phòng khám", uc13_categories, "Function")
build_sheet(wb, "UC1.4 - Thiết lập giá dịch vụ", "Thiết lập giá dịch vụ khám chữa bệnh", uc14_categories, "Function")

output_path = "G:/Project/Kiểm thử/TestCase-01-UC.xlsx"
wb.save(output_path)
print("Done creating TestCase-01-UC.xlsx!")
