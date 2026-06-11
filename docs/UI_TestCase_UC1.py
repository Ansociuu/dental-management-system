from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import math

wb = Workbook()

# Color constants
FILL_TITLE = PatternFill("solid", fgColor="99CCFF")
FILL_USECASE_HEADER = PatternFill("solid", fgColor="E2EFD9")
FILL_COL_HEADER = PatternFill("solid", fgColor="CCFFFF")
FILL_ROW_ALT = PatternFill("solid", fgColor="EFEFEF")
FILL_CATEGORY = PatternFill("solid", fgColor="D9E1F2")
FILL_NONE = PatternFill("solid", fgColor="FFFFFF")

FONT_TITLE = Font(name="Arial", bold=True, size=12)
FONT_HEADER = Font(name="Arial", bold=True, size=10)
FONT_BOLD = Font(name="Arial", bold=True, size=10)
FONT_NORMAL = Font(name="Arial", size=10)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def set_cell(ws, row, col, value, font=None, fill=None, align=None, border=True):
    cell = ws.cell(row=row, column=col)
    cell.value = value
    if font: cell.font = font
    if fill: cell.fill = fill
    if align: cell.alignment = align
    if border: cell.border = thin_border
    return cell

def apply_style(cell, fill, align, font=FONT_NORMAL):
    cell.fill = fill
    cell.alignment = align
    cell.border = thin_border
    cell.font = font

def apply_border_to_range(ws, start_row, end_row, start_col, end_col):
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            ws.cell(row=r, column=c).border = thin_border

def calculate_row_height(texts_with_widths, font_size=10):
    max_lines = 1
    for text, width in texts_with_widths:
        if not text: continue
        chars_per_line = max(1, width * 1.1)
        lines = len(str(text)) / chars_per_line
        lines += str(text).count('\n')
        if lines > max_lines:
            max_lines = lines
    return max(15, math.ceil(max_lines) * (font_size + 4))

def build_sheet(wb, sheet_name, use_case_title, categories_data, sheet_type="UI"):
    ws = wb.create_sheet(sheet_name)
    
    col_widths = [2, 18, 17, 32, 30, 28, 18, 35, 33, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    dv = DataValidation(type="list", formula1='"Passed,Failed,Not Run,Not Completed"', allow_blank=False)
    ws.add_data_validation(dv)

    ws.merge_cells("A1:J1")
    set_cell(ws, 1, 1, f"{sheet_type} Test Cases", FONT_TITLE, FILL_TITLE, ALIGN_CENTER)
    apply_border_to_range(ws, 1, 1, 1, 10)
    ws.row_dimensions[1].height = 22

    ws.merge_cells("A2:G3")
    set_cell(ws, 2, 1, f"Use Case: {use_case_title}", FONT_BOLD, FILL_USECASE_HEADER, ALIGN_CENTER)
    apply_border_to_range(ws, 2, 3, 1, 7)

    set_cell(ws, 2, 8, "Passed", FONT_NORMAL, FILL_NONE, ALIGN_CENTER)
    set_cell(ws, 3, 8, "Failed", FONT_NORMAL, FILL_NONE, ALIGN_CENTER)
    set_cell(ws, 4, 8, "Not Run", FONT_NORMAL, FILL_NONE, ALIGN_CENTER)
    set_cell(ws, 5, 8, "Not Completed", FONT_NORMAL, FILL_NONE, ALIGN_CENTER)
    set_cell(ws, 6, 8, "Number of test cases", FONT_NORMAL, FILL_NONE, ALIGN_CENTER)

    ws.merge_cells("B7:B8")
    set_cell(ws, 7, 2, "Category", FONT_HEADER, FILL_COL_HEADER, ALIGN_CENTER)
    ws.merge_cells("C7:C8")
    set_cell(ws, 7, 3, "Test Case ID", FONT_HEADER, FILL_COL_HEADER, ALIGN_CENTER)
    ws.merge_cells("D7:D8")
    set_cell(ws, 7, 4, "Test Case Description", FONT_HEADER, FILL_COL_HEADER, ALIGN_CENTER)
    ws.merge_cells("E7:F7")
    set_cell(ws, 7, 5, "Test Procedures", FONT_HEADER, FILL_COL_HEADER, ALIGN_CENTER)
    ws.merge_cells("G7:G8")
    set_cell(ws, 7, 7, "Test Input Value", FONT_HEADER, FILL_COL_HEADER, ALIGN_CENTER)
    ws.merge_cells("H7:H8")
    set_cell(ws, 7, 8, "Test Case Expected Result", FONT_HEADER, FILL_COL_HEADER, ALIGN_CENTER)
    ws.merge_cells("I7:I8")
    set_cell(ws, 7, 9, "Test Case Program Result", FONT_HEADER, FILL_COL_HEADER, ALIGN_CENTER)
    ws.merge_cells("J7:J8")
    set_cell(ws, 7, 10, "Status", FONT_HEADER, FILL_COL_HEADER, ALIGN_CENTER)

    set_cell(ws, 8, 5, "Steps to Perform", FONT_HEADER, FILL_COL_HEADER, ALIGN_LEFT)
    set_cell(ws, 8, 6, "Step Expected Result", FONT_HEADER, FILL_COL_HEADER, ALIGN_LEFT)

    apply_border_to_range(ws, 7, 8, 2, 10)
    
    ws.row_dimensions[7].height = 25
    ws.row_dimensions[8].height = 25

    current_row = 9
    total_tcs = 0
    for category_name, test_cases in categories_data:
        cat_start = current_row
        for tc in test_cases:
            steps = tc.get("steps", [])
            if not steps:
                continue
                
            total_tcs += 1
            tc_id = tc["id"]
            tc_desc = tc["desc"]
            expected_result = tc["expected"]
            tc_start = current_row
            tc_fill = FILL_ROW_ALT if (total_tcs % 2 == 0) else FILL_NONE

            for i, (step_text, step_expected, input_val) in enumerate(steps):
                for col in range(2, 11):
                    cell = ws.cell(row=current_row, column=col)
                    apply_style(cell, tc_fill, ALIGN_CENTER)
                    
                if i == 0:
                    ws.cell(row=current_row, column=3).value = tc_id
                    ws.cell(row=current_row, column=4).value = tc_desc
                    ws.cell(row=current_row, column=4).alignment = ALIGN_LEFT
                    ws.cell(row=current_row, column=8).value = expected_result
                    ws.cell(row=current_row, column=8).alignment = ALIGN_LEFT
                    ws.cell(row=current_row, column=9).value = ""
                    ws.cell(row=current_row, column=10).value = "Not Run"
                    dv.add(ws.cell(row=current_row, column=10))

                ws.cell(row=current_row, column=5).value = step_text
                ws.cell(row=current_row, column=5).alignment = ALIGN_LEFT
                ws.cell(row=current_row, column=6).value = step_expected
                ws.cell(row=current_row, column=6).alignment = ALIGN_LEFT
                ws.cell(row=current_row, column=7).value = input_val

                row_texts = [
                    (step_text, 30),
                    (step_expected, 28),
                    (tc_desc if i == 0 else "", 32),
                    (expected_result if i == 0 else "", 35)
                ]
                ws.row_dimensions[current_row].height = calculate_row_height(row_texts)
                
                current_row += 1

            tc_end = current_row - 1
            for col in [3, 4, 8, 9, 10]:
                if tc_start != tc_end:
                    ws.merge_cells(start_row=tc_start, start_column=col, end_row=tc_end, end_column=col)
                
                align = ALIGN_LEFT if col in [4, 8] else ALIGN_CENTER
                for r in range(tc_start, tc_end + 1):
                    c = ws.cell(row=r, column=col)
                    c.border = thin_border
                    c.alignment = align

        if current_row > cat_start:
            ws.merge_cells(start_row=cat_start, start_column=2, end_row=current_row-1, end_column=2)
            c = ws.cell(row=cat_start, column=2)
            c.value = category_name
            c.font = FONT_BOLD
            c.fill = FILL_CATEGORY
            c.alignment = ALIGN_CENTER
            apply_border_to_range(ws, cat_start, current_row - 1, 2, 2)

    last_row = current_row - 1
    
    ws.merge_cells("H2:I2")
    set_cell(ws, 2, 10, f'=COUNTIF($J$9:$J${last_row},"Passed")', FONT_NORMAL, FILL_NONE, ALIGN_CENTER)
    ws.merge_cells("H3:I3")
    set_cell(ws, 3, 10, f'=COUNTIF($J$9:$J${last_row},"Failed")', FONT_NORMAL, FILL_NONE, ALIGN_CENTER)
    ws.merge_cells("H4:I4")
    set_cell(ws, 4, 10, f'=COUNTIF($J$9:$J${last_row},"Not Run")', FONT_NORMAL, FILL_NONE, ALIGN_CENTER)
    ws.merge_cells("H5:I5")
    set_cell(ws, 5, 10, f'=COUNTIF($J$9:$J${last_row},"Not Completed")', FONT_NORMAL, FILL_NONE, ALIGN_CENTER)
    ws.merge_cells("H6:I6")
    set_cell(ws, 6, 10, total_tcs, FONT_NORMAL, FILL_NONE, ALIGN_CENTER)
    
    apply_border_to_range(ws, 2, 6, 8, 10)
    ws.auto_filter.ref = f"B7:J{last_row}"
    ws.freeze_panes = "A9"
    return ws


# ===================== UC1 UI DATA =====================

# UC1.1 Quản lý người dùng - UI Test Cases
uc11_ui = [
    {
        "id": "UC1.1_UI_001",
        "desc": "Kiểm tra hiển thị giao diện danh sách người dùng và phân quyền.",
        "steps": [
            ("1. Đăng nhập hệ thống với tài khoản Admin", "Đăng nhập thành công, hiển thị trang dashboard", "admin / Matkhau@123"),
            ("2. Chọn menu \"Quản lý người dùng\"", "Hiển thị bảng danh sách người dùng, thanh tìm kiếm và nút \"Thêm tài khoản\"", ""),
            ("3. Kiểm tra các cột trong bảng danh sách", "Bảng hiển thị đủ các cột: STT, Tên đăng nhập, Họ tên, Vai trò, Số điện thoại, Email, Trạng thái, Thao tác", ""),
            ("4. Di chuột (Hover) qua nút \"Thêm tài khoản\"", "Nút đổi sang tông màu sáng hơn, xuất hiện hiệu ứng trỏ chuột pointer", "")
        ],
        "expected": "Giao diện danh sách người dùng hiển thị đúng thiết kế, font chữ Arial rõ nét và căn chỉnh hài hòa."
    },
    {
        "id": "UC1.1_UI_002",
        "desc": "Kiểm tra hiển thị form Thêm tài khoản và validation lỗi nhập liệu.",
        "steps": [
            ("1. Click chọn nút \"Thêm tài khoản\"", "Form popup modal hiển thị giữa màn hình với các ô nhập và nút Lưu/Hủy", ""),
            ("2. Click chọn nút \"Lưu\" khi chưa nhập bất kỳ thông tin nào", "Các trường bắt buộc chuyển sang màu viền đỏ cảnh báo, xuất hiện tooltip báo lỗi nhập liệu dưới mỗi ô", ""),
            ("3. Nhập mật khẩu yếu (ít hơn 8 ký tự)", "Hệ thống hiển thị thông điệp cảnh báo màu đỏ ngay dưới ô nhập mật khẩu về độ mạnh", "Mật khẩu: 12345")
        ],
        "expected": "Form modal hiển thị cân đối, các lỗi validation hiển thị trực quan dạng thông điệp đỏ dưới các trường tương ứng."
    },
    {
        "id": "UC1.1_UI_003",
        "desc": "Kiểm tra hiển thị màu sắc trạng thái (badge) của người dùng.",
        "steps": [
            ("1. Quan sát cột Trạng thái trên bảng danh sách tài khoản", "Mỗi tài khoản hiển thị một nhãn (badge) trạng thái", ""),
            ("2. Kiểm tra màu sắc nhãn ACTIVE", "Hiển thị nhãn màu xanh lá với chữ \"Đang hoạt động\"", ""),
            ("3. Kiểm tra màu sắc nhãn INACTIVE", "Hiển thị nhãn màu xám hoặc đỏ nhạt với chữ \"Đã khóa\"", "")
        ],
        "expected": "Nhãn trạng thái hiển thị rõ ràng, đẹp mắt, có độ tương phản cao giúp dễ dàng phân biệt trạng thái tài khoản."
    }
]

# UC1.2 Quản lý bác sĩ - UI Test Cases
uc12_ui = [
    {
        "id": "UC1.2_UI_001",
        "desc": "Kiểm tra hiển thị giao diện Quản lý bác sĩ và bộ lọc.",
        "steps": [
            ("1. Chọn menu \"Quản lý bác sĩ\" từ menu hệ thống", "Hiển thị danh sách bác sĩ, ô tìm kiếm nhanh, dropdown lọc theo Chuyên khoa và nút \"Thêm bác sĩ\"", ""),
            ("2. Kiểm tra các cột trong bảng danh sách bác sĩ", "Bảng hiển thị các cột: STT, Ảnh đại diện, Họ tên, Ngày sinh, Giới tính, Chuyên khoa, Số CCHN, Trạng thái, Thao tác", "")
        ],
        "expected": "Giao diện danh sách bác sĩ hiển thị đúng layout, responsive tốt trên các kích thước màn hình phổ biến."
    },
    {
        "id": "UC1.2_UI_002",
        "desc": "Kiểm tra validation upload file ảnh đại diện hoặc chứng chỉ bác sĩ.",
        "steps": [
            ("1. Click chọn nút \"Thêm bác sĩ\"", "Form popup thêm mới bác sĩ hiển thị", ""),
            ("2. Nhấp chọn tải lên tệp ảnh/chứng chỉ có dung lượng vượt quá 5MB hoặc sai định dạng", "Hệ thống từ chối tải tệp, hiển thị Toast message hoặc text báo lỗi màu đỏ ngay tại vùng upload", "Tải lên file: CV_Doctor.exe (hoặc ảnh 10MB)")
        ],
        "expected": "Vùng upload hiển thị viền đỏ đứt nét và xuất hiện thông báo lỗi: \"Định dạng file không hỗ trợ hoặc kích thước vượt quá 5MB!\""
    },
    {
        "id": "UC1.2_UI_003",
        "desc": "Kiểm tra hiển thị Modal xem thông tin chi tiết hồ sơ bác sĩ.",
        "steps": [
            ("1. Click vào tên bác sĩ trong bảng danh sách", "Modal hiển thị thông tin chi tiết hồ sơ bác sĩ", "BS. Trần Quốc Bảo"),
            ("2. Kiểm tra các tab thông tin (Thông tin hành chính, Bằng cấp/Kinh nghiệm, Lịch sử chỉnh sửa)", "Chuyển đổi mượt mà giữa các tab, thông tin hiển thị rõ ràng, font chữ đẹp", "")
        ],
        "expected": "Modal xem chi tiết bác sĩ hiển thị đẹp mắt dưới dạng Card, thông tin phân loại rõ ràng và có nút Đóng góc trên bên phải."
    }
]

# UC1.3 Quản lý danh mục dịch vụ - UI Test Cases
uc13_ui = [
    {
        "id": "UC1.3_UI_001",
        "desc": "Kiểm tra hiển thị giao diện Danh mục dịch vụ và định dạng tiền tệ.",
        "steps": [
            ("1. Chọn chức năng \"Quản lý danh mục dịch vụ\" từ menu quản trị", "Hiển thị danh sách dịch vụ hiện có của phòng khám", ""),
            ("2. Kiểm tra cột \"Giá dịch vụ\" trên bảng hiển thị", "Mức giá của các dịch vụ được định dạng tiền tệ rõ ràng (ví dụ: 500.000 VNĐ)", ""),
            ("3. Kiểm tra cột \"Thời gian thực hiện\"", "Hiển thị đúng dạng số phút (ví dụ: 30 phút, 60 phút)", "")
        ],
        "expected": "Bảng danh mục hiển thị chuyên nghiệp, căn lề phải cho cột Giá dịch vụ, căn giữa cho cột Thời gian và Trạng thái."
    },
    {
        "id": "UC1.3_UI_002",
        "desc": "Kiểm tra validation trực quan form thêm/sửa dịch vụ.",
        "steps": [
            ("1. Click chọn nút \"Thêm dịch vụ\"", "Form modal thêm dịch vụ hiển thị", ""),
            ("2. Nhập Giá dịch vụ là số âm hoặc để trống Tên dịch vụ", "Ô nhập hiển thị cảnh báo lỗi bằng đường viền đỏ nổi bật, nút Lưu bị disabled", "Giá: -150.000 / Tên: (Trống)"),
            ("3. Chọn Nhóm dịch vụ từ Dropdown", "Dropdown hiển thị đúng danh sách các nhóm dịch vụ (Răng sứ, Hàn răng, Nhổ răng...)", "")
        ],
        "expected": "Giao diện cảnh báo lỗi nhập liệu ngay lập tức, hướng dẫn người dùng sửa giá trị không hợp lệ."
    }
]

# UC1.4 Thiết lập giá dịch vụ - UI Test Cases
uc14_ui = [
    {
        "id": "UC1.4_UI_001",
        "desc": "Kiểm tra hiển thị giao diện Thiết lập giá dịch vụ và lịch sử giá.",
        "steps": [
            ("1. Truy cập chức năng \"Thiết lập giá dịch vụ\"", "Hiển thị danh sách các dịch vụ kèm cột Giá hiện tại, Ngày áp dụng và icon xem lịch sử", ""),
            ("2. Click vào icon \"Xem lịch sử giá\" của một dịch vụ", "Giao diện mở ra bảng phụ hiển thị dòng thời gian (timeline) các lần thay đổi giá dịch vụ đó", "")
        ],
        "expected": "Bảng lịch sử giá hiển thị rõ ràng thông tin giá, ngày hiệu lực và người cập nhật theo dạng timeline đẹp mắt."
    },
    {
        "id": "UC1.4_UI_002",
        "desc": "Kiểm tra hiển thị Calendar/Date Picker khi chọn Ngày hiệu lực giá.",
        "steps": [
            ("1. Click cập nhật giá cho một dịch vụ", "Form thiết lập giá hiển thị", ""),
            ("2. Click vào trường \"Ngày hiệu lực\"", "Hộp thoại chọn ngày (Calendar Date Picker) hiển thị rõ ràng, dễ thao tác chọn ngày tháng", "")
        ],
        "expected": "Date picker hiển thị đúng tháng hiện tại, cho phép chuyển đổi tháng/năm dễ dàng và hiển thị ngày được chọn đúng định dạng DD/MM/YYYY."
    },
    {
        "id": "UC1.4_UI_003",
        "desc": "Kiểm tra giao diện cập nhật giá hàng loạt.",
        "steps": [
            ("1. Tick chọn nhiều dịch vụ trên bảng danh sách dịch vụ", "Các checkbox hiển thị trạng thái được tick chọn", ""),
            ("2. Click nút \"Cập nhật giá hàng loạt\"", "Modal popup cập nhật giá hàng loạt xuất hiện ở giữa màn hình", ""),
            ("3. Kiểm tra các ô nhập tỉ lệ tăng/giảm giá (%) và Ngày hiệu lực", "Các ô nhập hiển thị placeholder hướng dẫn rõ ràng", "")
        ],
        "expected": "Modal thiết lập giá hàng loạt hiển thị đúng layout, các nút chức năng \"Áp dụng\" và \"Hủy bỏ\" nổi bật."
    }
]

uc11_categories = [
    ("UC1.1 Giao diện người dùng", uc11_ui[:1]),
    ("UC1.1 Form & Trạng thái", uc11_ui[1:]),
]

uc12_categories = [
    ("UC1.2 Giao diện hồ sơ", uc12_ui[:1]),
    ("UC1.2 File Upload & Chi tiết", uc12_ui[1:]),
]

uc13_categories = [
    ("UC1.3 Danh mục & Giá", uc13_ui[:1]),
    ("UC1.3 Validation Form", uc13_ui[1:]),
]

uc14_categories = [
    ("UC1.4 Bảng giá & Lịch sử", uc14_ui[:1]),
    ("UC1.4 Date Picker & Hàng loạt", uc14_ui[1:]),
]


# Remove default sheet, build all sheets
if 'Sheet' in wb.sheetnames:
    del wb['Sheet']

# Build UI Test sheets for UC1
build_sheet(wb, "UC1.1 - UI Tài khoản", "Quản lý người dùng và phân quyền (UI)", uc11_categories, "UI")
build_sheet(wb, "UC1.2 - UI Bác sĩ", "Quản lý hồ sơ bác sĩ (UI)", uc12_categories, "UI")
build_sheet(wb, "UC1.3 - UI Danh mục dịch vụ", "Quản lý danh mục dịch vụ (UI)", uc13_categories, "UI")
build_sheet(wb, "UC1.4 - UI Thiết lập giá", "Thiết lập giá dịch vụ (UI)", uc14_categories, "UI")

output_path = "G:/Project/Kiểm thử/UI_TestCase_UC1.xlsx"
wb.save(output_path)
print("Done creating UI_TestCase_UC1.xlsx!")
