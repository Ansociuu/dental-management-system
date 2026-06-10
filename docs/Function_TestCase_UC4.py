from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import math

wb = Workbook()

# Color constants
FILL_TITLE = PatternFill("solid", fgColor="99CCFF")
FILL_USECASE_HEADER = PatternFill("solid", fgColor="E2EFD9")
FILL_COL_HEADER = PatternFill("solid", fgColor="CCFFFF")
FILL_ROW_ALT = PatternFill("solid", fgColor="EFEFEF")
FILL_CATEGORY = PatternFill("solid", fgColor="D9E1F2")
FILL_NONE = PatternFill("solid", fgColor="FFFFFF")

FONT_TITLE = Font(name="Arial", bold=True, size=12)
FONT_HEADER = Font(name="Arial", bold=True, size=10)
FONT_BOLD = Font(name="Arial", bold=True, size=10)
FONT_NORMAL = Font(name="Arial", size=10)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def set_cell(ws, row, col, value, font=None, fill=None, align=None, border=True):
    cell = ws.cell(row=row, column=col)
    cell.value = value
    if font: cell.font = font
    if fill: cell.fill = fill
    if align: cell.alignment = align
    if border: cell.border = thin_border
    return cell

def apply_style(cell, fill, align, font=FONT_NORMAL):
    cell.fill = fill
    cell.alignment = align
    cell.border = thin_border
    cell.font = font

def apply_border_to_range(ws, start_row, end_row, start_col, end_col):
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            ws.cell(row=r, column=c).border = thin_border

def calculate_row_height(texts_with_widths, font_size=10):
    max_lines = 1
    for text, width in texts_with_widths:
        if not text: continue
        chars_per_line = max(1, width * 1.1)
        lines = len(str(text)) / chars_per_line
        lines += str(text).count('\n')
        if lines > max_lines:
            max_lines = lines
    return max(15, math.ceil(max_lines) * (font_size + 4))

def build_sheet(wb, sheet_name, use_case_title, categories_data, sheet_type="Function"):
    ws = wb.create_sheet(sheet_name)
    
    col_widths = [2, 18, 17, 32, 30, 28, 18, 35, 33, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    dv = DataValidation(type="list", formula1='"Passed,Failed,Not Run,Not Completed"', allow_blank=False)
    ws.add_data_validation(dv)

    ws.merge_cells("A1:J1")
    set_cell(ws, 1, 1, f"{sheet_type} Test Cases", FONT_TITLE, FILL_TITLE, ALIGN_CENTER)
    apply_border_to_range(ws, 1, 1, 1, 10)
    ws.row_dimensions[1].height = 22

    ws.merge_cells("A2:G3")
    set_cell(ws, 2, 1, f"Use Case: {use_case_title}", FONT_BOLD, FILL_USECASE_HEADER, ALIGN_CENTER)
    apply_border_to_range(ws, 2, 3, 1, 7)

    set_cell(ws, 2, 8, "Passed", FONT_NORMAL, FILL_NONE, ALIGN_CENTER)
    set_cell(ws, 3, 8, "Failed", FONT_NORMAL, FILL_NONE, ALIGN_CENTER)
    set_cell(ws, 4, 8, "Not Run", FONT_NORMAL, FILL_NONE, ALIGN_CENTER)
    set_cell(ws, 5, 8, "Not Completed", FONT_NORMAL, FILL_NONE, ALIGN_CENTER)
    set_cell(ws, 6, 8, "Number of test cases", FONT_NORMAL, FILL_NONE, ALIGN_CENTER)

    ws.merge_cells("B7:B8")
    set_cell(ws, 7, 2, "Category", FONT_HEADER, FILL_COL_HEADER, ALIGN_CENTER)
    ws.merge_cells("C7:C8")
    set_cell(ws, 7, 3, "Test Case ID", FONT_HEADER, FILL_COL_HEADER, ALIGN_CENTER)
    ws.merge_cells("D7:D8")
    set_cell(ws, 7, 4, "Test Case Description", FONT_HEADER, FILL_COL_HEADER, ALIGN_CENTER)
    ws.merge_cells("E7:F7")
    set_cell(ws, 7, 5, "Test Procedures", FONT_HEADER, FILL_COL_HEADER, ALIGN_CENTER)
    ws.merge_cells("G7:G8")
    set_cell(ws, 7, 7, "Test Input Value", FONT_HEADER, FILL_COL_HEADER, ALIGN_CENTER)
    ws.merge_cells("H7:H8")
    set_cell(ws, 7, 8, "Test Case Expected Result", FONT_HEADER, FILL_COL_HEADER, ALIGN_CENTER)
    ws.merge_cells("I7:I8")
    set_cell(ws, 7, 9, "Test Case Program Result", FONT_HEADER, FILL_COL_HEADER, ALIGN_CENTER)
    ws.merge_cells("J7:J8")
    set_cell(ws, 7, 10, "Status", FONT_HEADER, FILL_COL_HEADER, ALIGN_CENTER)

    set_cell(ws, 8, 5, "Steps to Perform", FONT_HEADER, FILL_COL_HEADER, ALIGN_LEFT)
    set_cell(ws, 8, 6, "Step Expected Result", FONT_HEADER, FILL_COL_HEADER, ALIGN_LEFT)

    apply_border_to_range(ws, 7, 8, 2, 10)
    
    ws.row_dimensions[7].height = 25
    ws.row_dimensions[8].height = 25

    current_row = 9
    total_tcs = 0
    for category_name, test_cases in categories_data:
        cat_start = current_row
        for tc in test_cases:
            steps = tc.get("steps", [])
            if not steps:
                continue
                
            total_tcs += 1
            tc_id = tc["id"]
            tc_desc = tc["desc"]
            expected_result = tc["expected"]
            tc_start = current_row
            tc_fill = FILL_ROW_ALT if (total_tcs % 2 == 0) else FILL_NONE

            for i, (step_text, step_expected, input_val) in enumerate(steps):
                for col in range(2, 11):
                    cell = ws.cell(row=current_row, column=col)
                    apply_style(cell, tc_fill, ALIGN_CENTER)
                    
                if i == 0:
                    ws.cell(row=current_row, column=3).value = tc_id
                    ws.cell(row=current_row, column=4).value = tc_desc
                    ws.cell(row=current_row, column=4).alignment = ALIGN_LEFT
                    ws.cell(row=current_row, column=8).value = expected_result
                    ws.cell(row=current_row, column=8).alignment = ALIGN_LEFT
                    ws.cell(row=current_row, column=9).value = ""
                    ws.cell(row=current_row, column=10).value = "Not Run"
                    dv.add(ws.cell(row=current_row, column=10))

                ws.cell(row=current_row, column=5).value = step_text
                ws.cell(row=current_row, column=5).alignment = ALIGN_LEFT
                ws.cell(row=current_row, column=6).value = step_expected
                ws.cell(row=current_row, column=6).alignment = ALIGN_LEFT
                ws.cell(row=current_row, column=7).value = input_val

                row_texts = [
                    (step_text, 30),
                    (step_expected, 28),
                    (tc_desc if i == 0 else "", 32),
                    (expected_result if i == 0 else "", 35)
                ]
                ws.row_dimensions[current_row].height = calculate_row_height(row_texts)
                
                current_row += 1

            tc_end = current_row - 1
            for col in [3, 4, 8, 9, 10]:
                if tc_start != tc_end:
                    ws.merge_cells(start_row=tc_start, start_column=col, end_row=tc_end, end_column=col)
                
                align = ALIGN_LEFT if col in [4, 8] else ALIGN_CENTER
                for r in range(tc_start, tc_end + 1):
                    c = ws.cell(row=r, column=col)
                    c.border = thin_border
                    c.alignment = align

        if current_row > cat_start:
            ws.merge_cells(start_row=cat_start, start_column=2, end_row=current_row-1, end_column=2)
            c = ws.cell(row=cat_start, column=2)
            c.value = category_name
            c.font = FONT_BOLD
            c.fill = FILL_CATEGORY
            c.alignment = ALIGN_CENTER
            apply_border_to_range(ws, cat_start, current_row - 1, 2, 2)

    last_row = current_row - 1
    
    ws.merge_cells("H2:I2")
    set_cell(ws, 2, 10, f'=COUNTIF($J$9:$J${last_row},"Passed")', FONT_NORMAL, FILL_NONE, ALIGN_CENTER)
    ws.merge_cells("H3:I3")
    set_cell(ws, 3, 10, f'=COUNTIF($J$9:$J${last_row},"Failed")', FONT_NORMAL, FILL_NONE, ALIGN_CENTER)
    ws.merge_cells("H4:I4")
    set_cell(ws, 4, 10, f'=COUNTIF($J$9:$J${last_row},"Not Run")', FONT_NORMAL, FILL_NONE, ALIGN_CENTER)
    ws.merge_cells("H5:I5")
    set_cell(ws, 5, 10, f'=COUNTIF($J$9:$J${last_row},"Not Completed")', FONT_NORMAL, FILL_NONE, ALIGN_CENTER)
    ws.merge_cells("H6:I6")
    set_cell(ws, 6, 10, total_tcs, FONT_NORMAL, FILL_NONE, ALIGN_CENTER)
    
    apply_border_to_range(ws, 2, 6, 8, 10)
    ws.auto_filter.ref = f"B7:J{last_row}"
    ws.freeze_panes = "A9"
    return ws


# ===================== UC4 DATA =====================

# UC4.1 - Thiết lập mức tiền cơ bản cho một giờ
uc41_func = [
    {
        "id": "UC4.1_FUNC_001",
        "desc": "Thêm mới mức tiền cơ bản hợp lệ thành công.",
        "steps": [
            ("1. Đăng nhập hệ thống bằng tài khoản Quản lý", "Đăng nhập thành công, hiển thị trang quản trị", "quanly01 / Matkhau@123"),
            ("2. Truy cập chức năng \"Thiết lập mức tiền cơ bản\"", "Hiển thị danh sách mức tiền cơ bản đã thiết lập", ""),
            ("3. Chọn nút \"Thêm mới\"", "Hiển thị form nhập thông tin mức tiền cơ bản", ""),
            ("4. Nhập Mức tiền cơ bản, Ngày áp dụng và Ghi chú", "", "200.000 / Hôm nay / Áp dụng cho quý mới"),
            ("5. Nhấn nút \"Lưu\"", "Hệ thống thông báo lưu thành công và ghi log lịch sử", "")
        ],
        "expected": "Mức tiền cơ bản mới được lưu vào cơ sở dữ liệu với trạng thái ACTIVE và hiển thị trên danh sách."
    },
    {
        "id": "UC4.1_FUNC_002",
        "desc": "Chỉnh sửa thông tin mức tiền cơ bản chưa áp dụng.",
        "steps": [
            ("1. Truy cập danh sách mức tiền cơ bản", "Hiển thị danh sách", ""),
            ("2. Chọn một mức tiền cơ bản chưa đến ngày hiệu lực hoặc đang PENDING", "Chọn mức tiền cần sửa", ""),
            ("3. Chọn nút \"Chỉnh sửa\"", "Hiển thị thông tin mức tiền hiện tại trên form chỉnh sửa", ""),
            ("4. Nhập giá trị mới (Mức tiền, Ngày áp dụng) và nhấn \"Lưu\"", "", "220.000 / Ngày đầu tháng sau"),
            ("5. Nhấn \"Lưu\"", "Hệ thống báo cập nhật thành công và ghi nhận lịch sử thay đổi", "")
        ],
        "expected": "Dữ liệu được cập nhật mới trên hệ thống. Lịch sử thay đổi ghi nhận thông tin người sửa, thời gian, giá trị cũ và mới."
    },
    {
        "id": "UC4.1_FUNC_003",
        "desc": "Ngừng áp dụng mức tiền cơ bản (chuyển sang trạng thái INACTIVE).",
        "steps": [
            ("1. Truy cập danh sách mức tiền cơ bản", "Hiển thị danh sách", ""),
            ("2. Chọn một mức tiền cơ bản đang có hiệu lực (ACTIVE)", "Chọn mức tiền cần ngừng", ""),
            ("3. Nhấn nút \"Ngừng hiệu lực\" và xác nhận thao tác", "Hệ thống xác nhận thao tác thành công", "")
        ],
        "expected": "Trạng thái mức tiền cơ bản chuyển sang INACTIVE. Mức tiền này sẽ không được dùng để tính lương cho các ca làm việc phát sinh tiếp theo."
    },
    {
        "id": "UC4.1_FUNC_004",
        "desc": "Thiết lập mức tiền cơ bản có ngày áp dụng trong tương lai.",
        "steps": [
            ("1. Chọn \"Thêm mới\" tại màn hình thiết lập mức tiền cơ bản", "Hiển thị form nhập", ""),
            ("2. Nhập Mức tiền và Ngày áp dụng là ngày trong tương lai", "", "250.000 / Ngày 01 của tháng sau"),
            ("3. Nhấn nút \"Lưu\"", "Hệ thống lưu thành công mức tiền cơ bản mới", "")
        ],
        "expected": "Mức tiền cơ bản được lưu với trạng thái PENDING. Đến ngày hiệu lực, hệ thống sẽ tự động kích hoạt mức tiền này."
    },
    {
        "id": "UC4.1_FUNC_005",
        "desc": "Nhập mức tiền cơ bản nhỏ hơn hoặc bằng 0.",
        "steps": [
            ("1. Chọn \"Thêm mới\" mức tiền cơ bản", "Hiển thị form nhập", ""),
            ("2. Nhập mức tiền cơ bản nhỏ hơn hoặc bằng 0", "", "0 (hoặc -50.000)"),
            ("3. Nhấn nút \"Lưu\"", "Hệ thống báo lỗi và ngăn chặn hành vi lưu dữ liệu", "")
        ],
        "expected": "Hiển thị thông báo lỗi: \"Mức tiền cơ bản phải lớn hơn 0!\". Không có bản ghi nào được tạo trong database."
    },
    {
        "id": "UC4.1_FUNC_006",
        "desc": "Thiết lập mức tiền mới trùng khoảng thời gian hiệu lực với mức tiền đang có.",
        "steps": [
            ("1. Chọn \"Thêm mới\" mức tiền cơ bản", "Hiển thị form nhập", ""),
            ("2. Nhập khoảng ngày áp dụng trùng lắp với mức tiền cơ bản đang hoạt động (ACTIVE)", "", "200.000 / Trùng ngày với cấu hình hiện tại"),
            ("3. Nhấn nút \"Lưu\"", "Hệ thống cảnh báo lỗi trùng thời gian và từ chối lưu", "")
        ],
        "expected": "Hiển thị cảnh báo: \"Đã tồn tại mức tiền cơ bản khác trong cùng khoảng thời gian hiệu lực!\". Cấu hình trùng không được lưu."
    },
    {
        "id": "UC4.1_FUNC_007",
        "desc": "Người dùng không đủ thẩm quyền cố gắng thiết lập mức tiền cơ bản.",
        "steps": [
            ("1. Đăng nhập hệ thống với tài khoản Lễ tân hoặc Bác sĩ", "Đăng nhập thành công", "letan01 / Matkhau@123"),
            ("2. Cố gắng truy cập chức năng \"Thiết lập mức tiền cơ bản\" hoặc gọi API tương ứng", "Hệ thống báo lỗi phân quyền", "")
        ],
        "expected": "Hệ thống chặn truy cập và hiển thị thông báo: \"Bạn không có quyền quản lý tiền lương!\""
    }
]

uc41_categories = [
    ("UC4.1 Thêm & Sửa đổi", uc41_func[:2]),
    ("UC4.1 Trạng thái & Tương lai", uc41_func[2:4]),
    ("UC4.1 Ràng buộc & Phân quyền", uc41_func[4:]),
]

# UC4.2 - Thiết lập hệ số ca làm việc các ngày trong tuần
uc42_func = [
    {
        "id": "UC4.2_FUNC_001",
        "desc": "Thiết lập hệ số ca làm việc mới hợp lệ.",
        "steps": [
            ("1. Đăng nhập bằng tài khoản Quản lý", "Đăng nhập thành công, hiển thị trang quản trị", "quanly01 / Matkhau@123"),
            ("2. Truy cập chức năng \"Thiết lập hệ số ca làm việc\"", "Hiển thị danh sách cấu hình hệ số hiện tại", ""),
            ("3. Chọn nút \"Thêm mới\"", "Hiển thị form thiết lập hệ số", ""),
            ("4. Chọn Loại ngày, Ca làm việc, nhập Hệ số ca làm việc và Ngày áp dụng", "", "Chủ nhật / Ca tối / 1.5 / Hôm nay"),
            ("5. Nhấn \"Lưu\"", "Hệ thống xác thực dữ liệu thành công và thông báo lưu thành công", "")
        ],
        "expected": "Cấu hình hệ số ca làm việc mới được lưu vào cơ sở dữ liệu và hiển thị trên bảng danh sách cấu hình."
    },
    {
        "id": "UC4.2_FUNC_002",
        "desc": "Chỉnh sửa hệ số ca làm việc đang tồn tại.",
        "steps": [
            ("1. Chọn một cấu hình hệ số ca làm việc từ danh sách", "Chọn cấu hình", ""),
            ("2. Chọn nút \"Chỉnh sửa\"", "Hiển thị thông tin hệ số hiện tại trên form", ""),
            ("3. Cập nhật hệ số ca làm việc và nhấn \"Lưu\"", "", "Hệ số mới: 1.8"),
            ("4. Nhấn \"Lưu\"", "Hệ thống báo cập nhật thành công và ghi nhận lịch sử thay đổi", "")
        ],
        "expected": "Hệ số mới được cập nhật thành công. Lịch sử ghi nhận đúng giá trị cũ (1.5) và giá trị mới (1.8)."
    },
    {
        "id": "UC4.2_FUNC_003",
        "desc": "Sao chép cấu hình hệ số ca làm việc từ một cấu hình có sẵn.",
        "steps": [
            ("1. Chọn cấu hình hệ số ca làm việc trong danh sách", "Chọn cấu hình mẫu", ""),
            ("2. Chọn nút \"Sao chép\"", "Hiển thị form nhập với dữ liệu cũ được điền sẵn", ""),
            ("3. Thay đổi Loại ngày hoặc Ca làm việc và điều chỉnh hệ số phù hợp", "", "Loại ngày: Ngày lễ / Hệ số: 2.0"),
            ("4. Nhấn \"Lưu\"", "Hệ thống thông báo sao chép cấu hình thành công", "")
        ],
        "expected": "Một cấu hình hệ số ca làm việc mới được tạo thành công với các thông số đã chỉnh sửa."
    },
    {
        "id": "UC4.2_FUNC_004",
        "desc": "Ngừng hiệu lực cấu hình hệ số ca làm việc.",
        "steps": [
            ("1. Chọn cấu hình hệ số ca làm việc đang hoạt động", "Chọn cấu hình ACTIVE", ""),
            ("2. Nhấn nút \"Ngừng hiệu lực\"", "Hệ thống yêu cầu xác nhận thao tác", ""),
            ("3. Xác nhận", "Hệ thống thông báo cập nhật thành công", "")
        ],
        "expected": "Trạng thái cấu hình chuyển sang INACTIVE. Cấu hình này không được sử dụng để quy đổi giờ công kể từ thời điểm này."
    },
    {
        "id": "UC4.2_FUNC_005",
        "desc": "Thiết lập hệ số ca làm việc nhỏ hơn 1.0.",
        "steps": [
            ("1. Chọn \"Thêm mới\" hệ số ca làm việc", "Hiển thị form nhập", ""),
            ("2. Nhập thông tin hợp lệ nhưng điền Hệ số ca làm việc nhỏ hơn 1.0", "", "Hệ số: 0.8"),
            ("3. Nhấn \"Lưu\"", "Hệ thống phát hiện lỗi và từ chối lưu", "")
        ],
        "expected": "Hiển thị lỗi: \"Hệ số ca làm việc phải lớn hơn hoặc bằng 1.0!\". Bản ghi không được tạo."
    },
    {
        "id": "UC4.2_FUNC_006",
        "desc": "Thiết lập trùng lặp cấu hình hệ số ca làm việc tại cùng một thời điểm.",
        "steps": [
            ("1. Chọn \"Thêm mới\" hệ số ca làm việc", "Hiển thị form nhập", ""),
            ("2. Chọn trùng Loại ngày, Ca làm việc và Ngày áp dụng của một cấu hình ACTIVE khác đã có", "", "Ngày thường / Ca sáng / Hôm nay"),
            ("3. Nhấn \"Lưu\"", "Hệ thống báo lỗi trùng lặp cấu hình", "")
        ],
        "expected": "Hệ thống hiển thị cảnh báo: \"Đã tồn tại hệ số hiệu lực cho cặp Loại ngày và Ca làm việc này trong khoảng thời gian đã chọn!\". Từ chối lưu."
    }
]

uc42_categories = [
    ("UC4.2 Thêm & Sửa hệ số ca", uc42_func[:2]),
    ("UC4.2 Sao chép & Trạng thái", uc42_func[2:4]),
    ("UC4.2 Ràng buộc hệ số", uc42_func[4:]),
]

# UC4.3 - Nhập hệ số các ca cần xử lý phức tạp trong tháng
uc43_func = [
    {
        "id": "UC4.3_FUNC_001",
        "desc": "Nhập hệ số độ phức tạp cho ca khám đơn lẻ thành công.",
        "steps": [
            ("1. Đăng nhập bằng tài khoản Quản lý", "Đăng nhập thành công, hiển thị trang quản trị", "quanly01 / Matkhau@123"),
            ("2. Truy cập chức năng \"Nhập hệ số ca phức tạp\"", "Hiển thị danh sách ca khám trong tháng hiện tại", ""),
            ("3. Chọn một ca khám có trạng thái COMPLETED cần đánh giá", "Hiển thị chi tiết thông tin ca khám (Bác sĩ, bệnh nhân, dịch vụ)", "Bệnh nhân: Nguyễn Văn A"),
            ("4. Nhập giá trị hệ số độ phức tạp của ca khám", "", "Hệ số: 0.3 (Tiểu phẫu nhổ răng khôn)"),
            ("5. Nhấn nút \"Lưu\"", "Hệ thống thông báo lưu thành công", "")
        ],
        "expected": "Hệ số phức tạp được lưu vào cơ sở dữ liệu. Tổng hệ số bệnh nhân của ca trực tương ứng được tự động cập nhật cộng thêm 0.3."
    },
    {
        "id": "UC4.3_FUNC_002",
        "desc": "Nhập hàng loạt hệ số độ phức tạp cho nhiều ca khám cùng lúc.",
        "steps": [
            ("1. Tại màn hình nhập hệ số ca phức tạp, chọn nút \"Nhập hàng loạt\"", "Hiển thị danh sách các ca khám chưa được nhập hệ số", ""),
            ("2. Nhập hệ số độ phức tạp cho nhiều bệnh nhân trên bảng nhập dữ liệu", "", "Bệnh nhân B: 0.1, Bệnh nhân C: 0.4"),
            ("3. Nhấn nút \"Lưu tất cả\"", "Hệ thống thông báo cập nhật dữ liệu hàng loạt thành công", "")
        ],
        "expected": "Hệ thống lưu thành công toàn bộ hệ số độ phức tạp đã nhập cho các ca khám tương ứng."
    },
    {
        "id": "UC4.3_FUNC_003",
        "desc": "Chỉnh sửa hệ số độ phức tạp đã nhập trước đó.",
        "steps": [
            ("1. Chọn ca khám đã có hệ số độ phức tạp", "Hiển thị thông tin ca khám và hệ số hiện tại (0.3)", ""),
            ("2. Nhập hệ số mới và nhấn \"Lưu\"", "", "Hệ số mới: 0.5 (Cấy Implant)"),
            ("3. Nhấn \"Lưu\"", "Hệ thống báo cập nhật thành công và ghi log thay đổi", "")
        ],
        "expected": "Hệ thống cập nhật hệ số độ phức tạp thành 0.5. Lịch sử thay đổi được ghi nhận đầy đủ."
    },
    {
        "id": "UC4.3_FUNC_004",
        "desc": "Nhập hệ số độ phức tạp vượt quá giới hạn quy định.",
        "steps": [
            ("1. Chọn một ca khám COMPLETED", "Hiển thị form nhập", ""),
            ("2. Nhập hệ số độ phức tạp lớn hơn 0.5 hoặc nhỏ hơn 0", "", "Hệ số: 0.6 (hoặc -0.1)"),
            ("3. Nhấn nút \"Lưu\"", "Hệ thống phát hiện lỗi nhập liệu và ngăn chặn lưu", "")
        ],
        "expected": "Hiển thị cảnh báo lỗi: \"Hệ số độ phức tạp của một bệnh nhân phải nằm trong khoảng từ 0.0 đến 0.5!\". Cấu hình không được lưu."
    },
    {
        "id": "UC4.3_FUNC_005",
        "desc": "Cố gắng nhập hệ số phức tạp cho ca khám chưa hoàn tất khám bệnh.",
        "steps": [
            ("1. Tìm kiếm và chọn một ca khám đang ở trạng thái CHECKED_IN hoặc PENDING", "Thông tin ca khám hiển thị nhưng ô nhập hệ số bị disable hoặc hệ thống báo lỗi khi lưu", "Trạng thái: CHECKED_IN"),
            ("2. Thử nhập hệ số và nhấn \"Lưu\"", "Hệ thống báo lỗi trạng thái", "")
        ],
        "expected": "Hiển thị lỗi: \"Chỉ những ca khám ở trạng thái COMPLETED mới được phép nhập hệ số phức tạp!\". Không cho phép lưu."
    },
    {
        "id": "UC4.3_FUNC_006",
        "desc": "Cố gắng chỉnh sửa hệ số phức tạp của ca khám đã lập phiếu lương tháng.",
        "steps": [
            ("1. Chọn ca khám có hệ số độ phức tạp đã được đưa vào phiếu lương đã khóa/chốt", "Hiển thị chi tiết ca khám", ""),
            ("2. Thử thay đổi hệ số và nhấn \"Lưu\"", "Hệ thống từ chối chỉnh sửa", "")
        ],
        "expected": "Hệ thống hiển thị thông báo: \"Hệ số đã sử dụng để lập phiếu lương không được sửa trực tiếp. Vui lòng thực hiện quy trình điều chỉnh lương!\". Giao dịch bị hủy."
    }
]

uc43_categories = [
    ("UC4.3 Nhập & Cập nhật hệ số", uc43_func[:3]),
    ("UC4.3 Ràng buộc & Phiếu lương", uc43_func[3:]),
]

# UC4.4 - Lập phiếu lương cho một bác sĩ trong 1 tháng
uc44_func = [
    {
        "id": "UC4.4_FUNC_001",
        "desc": "Lập phiếu lương thành công cho một bác sĩ trong tháng hợp lệ.",
        "steps": [
            ("1. Đăng nhập hệ thống bằng tài khoản Quản lý", "Đăng nhập thành công, hiển thị trang quản trị", "quanly01 / Matkhau@123"),
            ("2. Truy cập chức năng \"Lập phiếu lương\"", "Hiển thị giao diện lập lương bác sĩ", ""),
            ("3. Chọn tên Bác sĩ và Tháng/Năm tính lương", "", "BS. Nguyễn Văn A / 05/2026"),
            ("4. Nhấn nút \"Tính lương\"", "Hệ thống tự động tổng hợp toàn bộ ca trực, số giờ làm việc, hệ số ca trực, hệ số ca phức tạp và mức tiền cơ bản hiệu lực để hiển thị bảng lương chi tiết", ""),
            ("5. Nhấn nút \"Lưu phiếu lương\"", "Hệ thống thông báo lập phiếu lương thành công", "")
        ],
        "expected": "Phiếu lương mới của bác sĩ Nguyễn Văn A trong tháng 05/2026 được tạo thành công với trạng thái PENDING_APPROVE/APPROVED. Dữ liệu lương lưu vào database."
    },
    {
        "id": "UC4.4_FUNC_002",
        "desc": "Cập nhật phụ cấp và ghi chú trên phiếu lương trước khi lưu.",
        "steps": [
            ("1. Chọn bác sĩ và tháng lập lương, nhấn \"Tính lương\"", "Hiển thị bảng tính lương chi tiết", "BS. Nguyễn Văn A / 05/2026"),
            ("2. Nhập số tiền Phụ cấp đặc thù phòng khám và nhập Ghi chú", "", "Phụ cấp: 1.000.000 / Ghi chú: Thưởng hiệu suất"),
            ("3. Hệ thống tự động cập nhật và tính lại tổng lương thực lĩnh", "Tổng lương cập nhật tăng thêm 1.000.000 VNĐ", ""),
            ("4. Nhấn nút \"Lưu phiếu lương\"", "Hệ thống lưu thành công phiếu lương đã chỉnh sửa", "")
        ],
        "expected": "Phiếu lương lưu đúng số tiền phụ cấp 1.000.000 VNĐ và nội dung ghi chú. Tổng lương thực lĩnh được tính chính xác."
    },
    {
        "id": "UC4.4_FUNC_003",
        "desc": "In và xuất file PDF phiếu lương của bác sĩ.",
        "steps": [
            ("1. Tìm kiếm và mở chi tiết phiếu lương đã lập của bác sĩ", "Hiển thị chi tiết phiếu lương", "PL-202605-001"),
            ("2. Chọn nút \"In phiếu lương\" hoặc \"Xuất PDF\"", "Hệ thống hiển thị bản xem trước in hoặc tạo file tải về", "")
        ],
        "expected": "Tải về thành công tệp tin PDF phiếu lương hiển thị đầy đủ, chính xác các thông tin chi tiết cấu thành lương và định dạng tiền tệ VNĐ."
    },
    {
        "id": "UC4.4_FUNC_004",
        "desc": "Lập lương khi để trống thông tin bắt buộc Bác sĩ hoặc Tháng.",
        "steps": [
            ("1. Truy cập giao diện lập phiếu lương", "Hiển thị giao diện", ""),
            ("2. Không chọn bác sĩ và nhấn nút \"Tính lương\"", "Hệ thống báo lỗi yêu cầu chọn bác sĩ", ""),
            ("3. Chọn bác sĩ nhưng không chọn tháng tính lương và nhấn \"Tính lương\"", "Hệ thống báo lỗi yêu cầu chọn tháng tính lương", "")
        ],
        "expected": "Hệ thống báo lỗi validate tương ứng và không cho phép thực hiện tính lương."
    },
    {
        "id": "UC4.4_FUNC_005",
        "desc": "Lập phiếu lương trong tháng bác sĩ không có bất kỳ ca trực nào.",
        "steps": [
            ("1. Chọn bác sĩ không đăng ký lịch trực và không có ca khám nào trong tháng cần tính lương", "", "BS. Lê Hoàng / 05/2026"),
            ("2. Nhấn nút \"Tính lương\"", "Hệ thống báo lỗi không có dữ liệu làm việc", "")
        ],
        "expected": "Hiển thị thông báo: \"Không có dữ liệu làm việc của bác sĩ trong tháng được chọn!\". Hệ thống không cho phép lưu phiếu lương rỗng."
    },
    {
        "id": "UC4.4_FUNC_006",
        "desc": "Cố gắng lập trùng phiếu lương cho một bác sĩ trong cùng một tháng.",
        "steps": [
            ("1. Chọn bác sĩ và tháng đã được lập phiếu lương trước đó", "", "BS. Nguyễn Văn A / 05/2026"),
            ("2. Nhấn nút \"Tính lương\"", "Hệ thống đưa ra cảnh báo trùng lặp phiếu lương", "")
        ],
        "expected": "Hiển thị thông báo: \"Phiếu lương cho bác sĩ này trong tháng đã chọn đã được lập trước đó! Không cho phép tạo trùng lặp!\". Nút Lưu bị khóa."
    }
]

uc44_categories = [
    ("UC4.4 Quy trình tính lương", uc44_func[:3]),
    ("UC4.4 Ràng buộc dữ liệu", uc44_func[3:]),
]

# UC4.5 - Báo cáo tiền lương tất cả bác sĩ trong 1 tháng
uc45_func = [
    {
        "id": "UC4.5_FUNC_001",
        "desc": "Xem báo cáo tổng hợp tiền lương tất cả bác sĩ trong tháng thành công.",
        "steps": [
            ("1. Đăng nhập hệ thống bằng tài khoản Quản lý", "Đăng nhập thành công, hiển thị trang quản trị", "quanly01 / Matkhau@123"),
            ("2. Truy cập chức năng \"Báo cáo lương tháng\"", "Hiển thị giao diện báo cáo lương tháng và bộ lọc", ""),
            ("3. Chọn Tháng và Năm báo cáo cần xem", "", "Tháng 5 / Năm 2026"),
            ("4. Nhấn nút \"Xem báo cáo\"", "Hệ thống truy xuất dữ liệu phiếu lương đã lập và hiển thị bảng báo cáo tổng hợp", "")
        ],
        "expected": "Báo cáo hiển thị đầy đủ danh sách bác sĩ, số ca làm, số giờ làm, tổng lương từng người. Tổng chi phí lương toàn phòng khám khớp với tổng các phiếu lương APPROVED/PAID."
    },
    {
        "id": "UC4.5_FUNC_002",
        "desc": "Tìm kiếm nhanh thông tin bác sĩ trên báo cáo lương tháng.",
        "steps": [
            ("1. Tại giao diện báo cáo lương tháng 5/2026 đang hiển thị", "Hiển thị bảng báo cáo đầy đủ", ""),
            ("2. Nhập từ khóa tên bác sĩ vào thanh tìm kiếm", "Bảng báo cáo lọc nhanh realtime danh sách bác sĩ tương ứng", "Trần Thị Bích"),
            ("3. Nhập từ khóa mã bác sĩ vào thanh tìm kiếm", "Hiển thị thông tin lương bác sĩ có mã tương ứng", "BS002")
        ],
        "expected": "Hệ thống thực hiện lọc kết quả chính xác, trả về dữ liệu nhanh chóng dưới 2 giây."
    },
    {
        "id": "UC4.5_FUNC_003",
        "desc": "Xuất báo cáo lương tháng ra file Excel hoặc PDF.",
        "steps": [
            ("1. Chọn tháng báo cáo và nhấn \"Xem báo cáo\"", "Bảng số liệu hiển thị đầy đủ", ""),
            ("2. Chọn nút \"Xuất Excel\" hoặc \"Xuất PDF\"", "Hệ thống tiến hành kết xuất dữ liệu và tải file về máy", "")
        ],
        "expected": "Tải về file Excel/PDF báo cáo lương tháng thành công, nội dung chính xác, định dạng VNĐ và cấu trúc bảng trùng khớp với hệ thống."
    },
    {
        "id": "UC4.5_FUNC_004",
        "desc": "Xem báo cáo lương tháng khi chưa chọn tháng báo cáo.",
        "steps": [
            ("1. Truy cập chức năng báo cáo lương tháng", "Hiển thị giao diện bộ lọc", ""),
            ("2. Để trống ô chọn tháng/năm báo cáo và nhấn \"Xem báo cáo\"", "Hệ thống cảnh báo lỗi", "")
        ],
        "expected": "Hiển thị thông báo yêu cầu chọn tháng/năm báo cáo. Không thực hiện truy vấn cơ sở dữ liệu."
    },
    {
        "id": "UC4.5_FUNC_005",
        "desc": "Xem báo cáo lương của tháng chưa có dữ liệu tính lương.",
        "steps": [
            ("1. Chọn tháng/năm chưa lập bất kỳ phiếu lương nào", "", "Tháng 12 / Năm 2026"),
            ("2. Nhấn nút \"Xem báo cáo\"", "Hệ thống thông báo không có dữ liệu lương", "")
        ],
        "expected": "Hiển thị thông báo: \"Không có dữ liệu lương trong tháng được chọn!\". Bảng thống kê hiển thị trống."
    }
]

uc45_categories = [
    ("UC4.5 Tổng hợp lương tháng", uc45_func[:3]),
    ("UC4.5 Bộ lọc & Ngoại lệ", uc45_func[3:]),
]

# UC4.6 - Báo cáo tiền lương của một bác sĩ trong một năm
uc46_func = [
    {
        "id": "UC4.6_FUNC_001",
        "desc": "Xem báo cáo tiền lương của một bác sĩ trong một năm thành công.",
        "steps": [
            ("1. Đăng nhập hệ thống bằng tài khoản Quản lý", "Đăng nhập thành công, hiển thị trang quản trị", "quanly01 / Matkhau@123"),
            ("2. Truy cập chức năng \"Báo cáo lương năm theo bác sĩ\"", "Hiển thị giao diện tìm kiếm và chọn năm", ""),
            ("3. Chọn Bác sĩ và Năm báo cáo cần xem", "", "BS. Nguyễn Văn A / Năm 2026"),
            ("4. Nhấn nút \"Xem báo cáo\"", "Hệ thống truy xuất dữ liệu lương của bác sĩ theo từng tháng trong năm và hiển thị bảng chi tiết", "")
        ],
        "expected": "Báo cáo hiển thị chi tiết tiền lương qua 12 tháng của bác sĩ Nguyễn Văn A. Tổng số ca, tổng giờ, tổng phụ cấp, tổng lương cả năm được cộng dồn chính xác."
    },
    {
        "id": "UC4.6_FUNC_002",
        "desc": "Xem chi tiết lương của một tháng cụ thể từ báo cáo năm.",
        "steps": [
            ("1. Thực hiện xem báo cáo lương năm của bác sĩ Nguyễn Văn A", "Hiển thị bảng lương 12 tháng", ""),
            ("2. Click chọn vào dòng lương của \"Tháng 5\" trên bảng báo cáo", "Hệ thống mở giao diện chi tiết hoặc popover hiển thị chi tiết phiếu lương tháng 5", "")
        ],
        "expected": "Hiển thị đầy đủ thông tin chi tiết cấu thành lương tháng 5 của bác sĩ bao gồm: danh sách ca làm việc, hệ số ca trực, hệ số phức tạp, và phụ cấp."
    },
    {
        "id": "UC4.6_FUNC_003",
        "desc": "Xuất báo cáo lương năm của bác sĩ ra file Excel/PDF.",
        "steps": [
            ("1. Xem báo cáo lương năm của bác sĩ", "Báo cáo hiển thị trên màn hình", "BS. Nguyễn Văn A / 2026"),
            ("2. Nhấn nút \"Xuất Excel\" hoặc \"Xuất PDF\"", "Hệ thống thực hiện tải file báo cáo về máy", "")
        ],
        "expected": "File báo cáo được tải về thành công với định dạng và nội dung hiển thị chính xác lương năm của bác sĩ được chọn."
    },
    {
        "id": "UC4.6_FUNC_004",
        "desc": "Xem báo cáo lương năm khi chưa chọn Bác sĩ hoặc chọn Năm.",
        "steps": [
            ("1. Truy cập chức năng báo cáo lương năm theo bác sĩ", "Hiển thị giao diện bộ lọc", ""),
            ("2. Để trống ô chọn bác sĩ (hoặc chọn năm), nhấn \"Xem báo cáo\"", "Hệ thống đưa ra cảnh báo lỗi", "")
        ],
        "expected": "Hệ thống báo lỗi validate yêu cầu nhập đầy đủ bác sĩ và năm báo cáo để thực hiện truy vấn."
    },
    {
        "id": "UC4.6_FUNC_005",
        "desc": "Xem báo cáo lương năm của bác sĩ chưa từng được lập lương trong năm.",
        "steps": [
            ("1. Chọn bác sĩ mới chưa phát sinh phiếu lương nào trong năm báo cáo", "", "BS. Đỗ Minh / Năm 2026"),
            ("2. Nhấn nút \"Xem báo cáo\"", "Hệ thống thông báo không có dữ liệu lương", "")
        ],
        "expected": "Hiển thị thông báo: \"Không có dữ liệu lương của bác sĩ trong năm đã chọn!\". Giao diện bảng hiển thị rỗng."
    }
]

uc46_categories = [
    ("UC4.6 Thống kê lương năm bác sĩ", uc46_func[:3]),
    ("UC4.6 Bộ lọc & Ngoại lệ", uc46_func[3:]),
]

# UC4.7 - Báo cáo tiền lương tất cả bác sĩ trong 1 năm
uc47_func = [
    {
        "id": "UC4.7_FUNC_001",
        "desc": "Xem báo cáo tiền lương tất cả bác sĩ trong một năm thành công.",
        "steps": [
            ("1. Đăng nhập hệ thống bằng tài khoản Quản lý", "Đăng nhập thành công, hiển thị trang quản trị", "quanly01 / Matkhau@123"),
            ("2. Truy cập chức năng \"Báo cáo lương năm toàn bộ bác sĩ\"", "Hiển thị giao diện báo cáo lương năm", ""),
            ("3. Chọn Năm báo cáo và nhấn \"Xem báo cáo\"", "", "Năm 2026"),
            ("4. Nhấn nút \"Xem báo cáo\"", "Hệ thống tổng hợp và hiển thị bảng lương năm của tất cả bác sĩ", "")
        ],
        "expected": "Báo cáo hiển thị đầy đủ danh sách bác sĩ có phát sinh lương, tổng ca làm, tổng giờ làm, tổng lương của từng người trong năm. Tổng lương toàn bộ bác sĩ hiển thị chính xác."
    },
    {
        "id": "UC4.7_FUNC_002",
        "desc": "Sử dụng bộ lọc tên bác sĩ trên báo cáo lương năm toàn bộ bác sĩ.",
        "steps": [
            ("1. Tại giao diện báo cáo lương năm toàn bộ bác sĩ đang hiển thị", "Hiển thị bảng báo cáo đầy đủ", ""),
            ("2. Nhập từ khóa tên bác sĩ cần lọc vào ô tìm kiếm", "Bảng báo cáo tự động ẩn các bác sĩ khác và chỉ hiển thị bác sĩ khớp từ khóa", "Nguyễn Văn A")
        ],
        "expected": "Hệ thống lọc chính xác và nhanh chóng. Chỉ hiển thị thông tin bác sĩ Nguyễn Văn A."
    },
    {
        "id": "UC4.7_FUNC_003",
        "desc": "Xuất báo cáo lương năm của toàn bộ bác sĩ ra file Excel/PDF.",
        "steps": [
            ("1. Xem báo cáo lương năm toàn bộ bác sĩ", "Bảng báo cáo hiển thị đầy đủ", ""),
            ("2. Nhấn nút \"Xuất Excel\" hoặc \"Xuất PDF\"", "Hệ thống tiến hành tải file về máy", "")
        ],
        "expected": "Tải file báo cáo về máy thành công, hiển thị đầy đủ danh sách bác sĩ và số liệu tài chính lương cả năm theo định dạng VNĐ."
    },
    {
        "id": "UC4.7_FUNC_004",
        "desc": "Xem báo cáo lương năm toàn bộ bác sĩ trong năm không có dữ liệu lương.",
        "steps": [
            ("1. Chọn năm tương lai chưa phát sinh bất kỳ phiếu lương nào", "", "Năm 2030"),
            ("2. Nhấn nút \"Xem báo cáo\"", "Hệ thống báo không có dữ liệu lương", "")
        ],
        "expected": "Hiển thị thông báo: \"Không có dữ liệu lương trong năm đã chọn!\". Bảng thống kê hiển thị trống."
    }
]

uc47_categories = [
    ("UC4.7 Tổng hợp lương năm", uc47_func[:3]),
    ("UC4.7 Bộ lọc & Ngoại lệ", uc47_func[3:]),
]

# Remove default sheet, build all sheets
if 'Sheet' in wb.sheetnames:
    del wb['Sheet']

# Build Functional Test sheets
build_sheet(wb, "UC4.1 - Mức tiền cơ bản", "Thiết lập mức tiền cơ bản cho một giờ", uc41_categories, "Function")
build_sheet(wb, "UC4.2 - Hệ số ca làm việc", "Thiết lập hệ số ca làm việc các ngày trong tuần", uc42_categories, "Function")
build_sheet(wb, "UC4.3 - Hệ số ca phức tạp", "Nhập hệ số các ca cần xử lý phức tạp trong tháng", uc43_categories, "Function")
build_sheet(wb, "UC4.4 - Lập phiếu lương", "Lập phiếu lương cho một bác sĩ trong 1 tháng", uc44_categories, "Function")
build_sheet(wb, "UC4.5 - Báo cáo lương tháng", "Báo cáo tiền lương tất cả bác sĩ trong 1 tháng", uc45_categories, "Function")
build_sheet(wb, "UC4.6 - Lương năm một bác sĩ", "Báo cáo tiền lương của một bác sĩ trong một năm", uc46_categories, "Function")
build_sheet(wb, "UC4.7 - Lương năm các bác sĩ", "Báo cáo tiền lương tất cả bác sĩ trong 1 năm", uc47_categories, "Function")

output_path = "G:/Project/Kiểm thử/TestCase-04-UC.xlsx"
wb.save(output_path)
print("Done!")
